from django.shortcuts import render, redirect, get_object_or_404   
from django.contrib.auth.decorators import login_required, user_passes_test   
from django.contrib.auth import login, authenticate   
from django.contrib import messages   
from django.http import HttpResponseForbidden, HttpResponseRedirect, JsonResponse, HttpResponse   
from django.views.decorators.csrf import csrf_protect   
from django.core.paginator import Paginator   
from rest_framework.decorators import api_view   
from rest_framework.response import Response   
from django.template.loader import get_template
from django.http import JsonResponse
from transporter_app.models import City, Pickup
import json
from django.db import transaction
from io import BytesIO   
from xhtml2pdf import pisa   
from functools import wraps  
from .models import BookingType, CNotes, DeliveryDestination, Dealer, CNote, Article, ArtType
import logging   
from django.http import JsonResponse
from transporter_app.models import Transporter
from django.db import connection
from django.views.decorators.http import require_POST
from django.shortcuts import render, get_object_or_404
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login as auth_login
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import CNote  # ‡§Ø‡§¶‡§ø ‡§Ü‡§™ CNote ‡§Æ‡•â‡§°‡§≤ ‡§ï‡§æ ‡§â‡§™‡§Ø‡•ã‡§ó ‡§ï‡§∞ ‡§∞‡§π‡•á ‡§π‡•à‡§Ç
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import CNote
from transporter_app.models import Pickup  # Import Pickup from transporter_app
from django.urls import reverse
from django.db.models import Count
from django.http import JsonResponse
from .models import CNote
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum
#from .models import Transporter, Pickup  # Assuming you have these models
from django.shortcuts import render, get_object_or_404
from .models import Dealer, CNote  # Make sure these models are correctly imported
from transporter_app.models import Transporter, Pickup  # Import Pickup from transporter_app
from django.contrib.auth import authenticate, login as auth_login
from django.shortcuts import redirect, render
from django.contrib import messages
from transporter_app.models import Transporter
from transporter_app.models import Transporter  # Make sure this import is correct
import logging
from django.http import JsonResponse
from transporter_app.models import City  # ‡§∏‡§π‡•Ä ‡§Ü‡§Ø‡§æ‡§§
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import CNote  # Ensure your models are imported correctly
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.utils import timezone
from .models import CNote
from .models import CNotes
from transporter_app.models import Pickup
import logging
from transporter_app.models import Transporter  # Transporter ‡§Æ‡•â‡§°‡§≤ ‡§ï‡§æ ‡§∏‡§π‡•Ä ‡§Ü‡§Ø‡§æ‡§§
from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from xml.etree.ElementTree import Element, SubElement, tostring
from .models import BookingType, DeliveryDestination, Dealer, CNote, Article, ArtType, LoadingSheetSummary, LoadingSheetDetail
import logging
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from .models import CNote, LoadingSheetSummary, LoadingSheetDetail
from transporter_app.models import Transporter
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import Http404
from .models import CNotes
from django.shortcuts import get_object_or_404
from django.db import connection
from django.http import JsonResponse
from transporter_app.models import PartyMaster
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
import json
import logging
from django.contrib import messages
from django.shortcuts import redirect
from django.views.decorators.http import require_GET, require_POST
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import CNotes, DeliveryDestination, Article, ArtType
from decimal import Decimal
import openpyxl
from django.http import HttpResponse
from transporter_app.models import DDMDetails
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, F, FloatField, Value
from django.db.models.functions import Cast, Coalesce, NullIf
import xlsxwriter
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Q
from collections import defaultdict
from django.contrib.postgres.aggregates import StringAgg
import platform
if platform.system() == "Windows":
    import win32print
else:
    win32print = None
import win32ui
from PIL import Image, ImageWin
from io import BytesIO
import base64
from PIL import Image, ImageWin, ImageFont, ImageDraw
import qrcode


# Set up logging   
logger = logging.getLogger(__name__)   

@login_required
def render_to_pdf(template_src, context_dict):   
    template = get_template(template_src)   
    html = template.render(context_dict)   
    result = BytesIO()   
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)   
    if not pdf.err:   
        return result.getvalue()   
    return None 
  
def user_owns_cnote(view_func):
    @wraps(view_func)
    def wrapper(request, cnote_id, *args, **kwargs):
        # Get the CNote object using the provided cnote_id
        cnote = get_object_or_404(CNote, id=cnote_id)
        
        # Check if the dealer associated with the CNote is the same as the logged-in user
        if cnote.dealer.user != request.user:
            return HttpResponseForbidden("You do not have permission to access this CNote.")
        
        # Call the original view function
        return view_func(request, cnote_id, *args, **kwargs)
    
    return wrapper


logger = logging.getLogger(__name__)


def custom_login_redirect(request):

    if request.method == 'POST':

        user_type = request.POST.get('user_type')

        username = request.POST.get('username')

        password = request.POST.get('password')

        

        logger.info(f"Login attempt - User Type: {user_type}, Username: {username}")

        

        # First, check if the user exists and credentials are correct

        user = authenticate(request, username=username, password=password)

        

        if user is not None:

            logger.info(f"User  authenticated successfully: {user.username}")

            

            # Debug query to check user flags

            with connection.cursor() as cursor:

                cursor.execute("""

                    SELECT is_dealer, is_transporter, is_active, is_staff 

                    FROM dealer_app_customuser 

                    WHERE username = %s

                """, [username])

                row = cursor.fetchone()

                if row:

                    logger.info(f"User  flags - is_dealer: {row[0]}, is_transporter: {row[1]}, is_active: {row[2]}, is_staff: {row[3]}")

                else:

                    logger.error(f"No user flags found for username: {username}")


            if user_type == 'dealer':

                # Check if user is marked as dealer in CustomUser  model

                if not user.is_dealer:

                    logger.warning(f"User  {username} attempted dealer login but is_dealer=False")

                    messages.error(request, 'Your account is not authorized as a dealer.')

                    return render(request, 'registration/login.html')

                

                # Check if there's an associated Dealer record

                try:

                    dealer = Dealer.objects.select_related('user').get(user=user)

                    logger.info(f"Found dealer record for user {username}: {dealer.name}")

                    login(request, user)

                    logger.info(f"Dealer login successful: {username}")

                    return redirect('dealer:create_cnotes')

                except Dealer.DoesNotExist:

                    logger.error(f"No Dealer record found for user {username} despite is_dealer=True")

                    messages.error(request, 'Dealer account not properly configured. Please contact support.')

                    return render(request, 'registration/login.html')

                

            elif user_type == 'transporter':

                # Check if user is marked as transporter in CustomUser  model

                if not user.is_transporter:

                    logger.warning(f"User  {username} attempted transporter login but is_transporter=False")

                    messages.error(request, 'Your account is not authorized as a transporter.')

                    return render(request, 'registration/login.html')

                

                # Check if there's an associated Transporter record

                try:

                    transporter = Transporter.objects.select_related('user').get(user=user)

                    logger.info(f"Found transporter record for user {username}")

                    login(request, user)

                    logger.info(f"Transporter login successful: {username}")

                    return redirect('transporter:home')

                except Transporter.DoesNotExist:

                    logger.error(f"No Transporter record found for user {username} despite is_transporter=True")

                    messages.error(request, 'Transporter account not properly configured. Please contact support.')

                    return render(request, 'registration/login.html')

            else:

                logger.warning(f"Invalid user type specified: {user_type}")

                messages.error(request, 'Invalid account type selected.')

        else:

            logger.warning(f"Authentication failed for username: {username}")

            messages.error(request, 'Invalid username or password.')


    return render(request, 'registration/login.html')    
@login_required
def login_redirect(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user_type = request.POST.get('user_type')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            print("Authenticated user:", user.username)  # Debug statement
            # Redirect based on user type
            if user_type == 'dealer':
                return redirect('dealer:create_cnotes')  # Redirect to dealer page
            elif user_type == 'transporter':
                return redirect('transporter:home')  # Use the namespace here
        else:
            print("Authentication failed for user:", username)  # Debugging output
            messages.error(request, 'Invalid username or password.')

    return render(request, 'registration/login.html')

@login_required  
@user_owns_cnote  
def get_cnote(request, cnote_id):  
    try:  
        cnote = CNote.objects.get(id=cnote_id)  
        articles = Article.objects.filter(cnote=cnote)  
        cnote_data = {  
            'id': cnote.id,  
            'booking_type': cnote.booking_type,  
            'delivery_destination': cnote.delivery_destination,  
            'consignee_name': cnote.consignee_name,  
            'payment_type': cnote.payment_type,  
            'grand_total': cnote.grand_total,  
            'created_at': cnote.created_at.strftime('%Y-%m-%d %H:%M:%S'),  
            'articles': [  
                {  
                    'article_type': article.article_type,  
                    'art': article.art,  
                    'art_type': article.art_type,  
                    'said_to_cont': article .said_to_cont,  
                    'art_amt': article.art_amt  
                } for article in articles  
            ]  
        }  
        return JsonResponse({'status': 'success', 'data': cnote_data})  
    except CNote.DoesNotExist:  
        logger.warning(f"CNote not found: {cnote_id}")  
        return JsonResponse({'status': 'error', 'message': 'CNote not found'}, status=404)  
    except Exception as e:  
        logger.error(f"Error getting CNote: {str(e)}")  
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)  

@login_required  
def list_cnotes(request):  
    try:  
        cnotes = CNote.objects.filter(dealer__user=request.user).order_by('-created_at')  
        paginator = Paginator(cnotes, 10)  
        page_number = request.GET.get('page')  
        page_obj = paginator.get_page(page_number)  
        cnotes_data = [  
            {  
                'id': cnote.id,  
                'booking_type': cnote.booking_type,  
                'delivery_destination': cnote.delivery_destination,  
                'consignee_name': cnote.consignee_name,  
                'payment_type': cnote.payment_type,  
                'grand_total': cnote.grand_total,  
                'created_at': cnote.created_at.strftime('%Y-%m-%d %H:%M:%S')  
            } for cnote in page_obj  
        ]  
        return JsonResponse({'status': 'success', 'data': cnotes_data, 'count': paginator.count})  
    except Exception as e:  
        logger.error(f"Error getting CNotes: {str(e)}")  
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)  

@login_required  
@user_owns_cnote  
def update_cnote(request, cnote_id):  
    cnote = get_object_or_404(CNote, id=cnote_id) 
    if request.method == 'POST':  
        # Access form data using request.POST  
        dealer_id = request.POST.get('dealer_id')  
        booking_type = request.POST.get('booking_type')  
        delivery_type = request.POST.get('delivery_type')  
        delivery_method = request.POST.get('delivery_method')  
        delivery_destination = request.POST.get('delivery_destination')  
        ewayBillNumber = request.POST.get('ewayBillNumber')  
        consignor_name = request.POST.get('consignor_name')  
        consignor_mobile = request.POST.get('consignor_mobile')  
        consignor_gst = request.POST.get('consignor_gst')  
        consignor_address = request.POST.get('consignor_address')  
        consignee_name = request.POST.get('consignee_name')  
        consignee_mobile = request.POST.get('consignee_mobile')  
        consignee_gst = request.POST.get('consignee_gst')  
        consignee_address = request.POST.get('consignee_address')  
        article = request.POST.get('article')  
        art = request.POST.get('art')  
        art_type = request.POST.get('art_type')  
        said_to_cont = request.POST.get('said_to_cont')  
        art_amt = request.POST.get('art_amt')  
        actual_weight = request.POST.get('actual_weight')  
        charged_weight = request.POST.get('charged_weight')  
        weight_rate = request.POST.get('weight_rate')  
        weight_amount = request.POST.get('weight_amount')  
        fix_amount = request.POST.get('fix_amount')  
        invoice_number = request.POST.get('invoice_number')  
        declared_value = request.POST.get('declared_value')  
        risk_type = request.POST.get('risk_type')  
        pod_required = request.POST.get('pod_required')  
        freight = request.POST.get('freight')  
        docket_charge = request.POST.get('docket_charge')  
        door_delivery_charge = request.POST.get('door_delivery_charge')  
        handling_charge = request.POST.get('handling_charge')  
        pickup_charge = request.POST.get('pickup_charge')  
        transhipment_charge = request.POST.get('transhipment_charge')  
        insurance = request.POST.get('insurance')  
        fuel_surcharge = request.POST.get('fuel_surcharge')  
        commission = request.POST.get('commission')  
        other_charge = request.POST.get('other_charge')  
        carrier_risk = request.POST.get('carrier_risk')  
        grand_total = request.POST.get('grand_total')  

        # Update the CNote object and save it to the database  
        cnote.dealer = Dealer.objects.get(id=dealer_id)  
        cnote.booking_type = booking_type  
        cnote.delivery_type = delivery_type  
        cnote.delivery_method = delivery_method  
        cnote .delivery_destination = delivery_destination  
        cnote.ewayBillNumber = ewayBillNumber  
        cnote.consignor_name = consignor_name  
        cnote.consignor_mobile = consignor_mobile  
        cnote.consignor_gst = consignor_gst  
        cnote.consignor_address = consignor_address  
        cnote.consignee_name = consignee_name  
        cnote.consignee_mobile = consignee_mobile  
        cnote.consignee_gst = consignee_gst  
        cnote.consignee_address = consignee_address  
        cnote.article = article  
        cnote.art = art  
        cnote.art_type = art_type  
        cnote.said_to_cont = said_to_cont  
        cnote.art_amt = art_amt  
        cnote.actual_weight = actual_weight  
        cnote.charged_weight = charged_weight  
        cnote.weight_rate = weight_rate  
        cnote.weight_amount = weight_amount  
        cnote.fix_amount = fix_amount  
        cnote.invoice_number = invoice_number  
        cnote.declared_value = declared_value  
        cnote.risk_type = risk_type  
        cnote.pod_required = pod_required  
        cnote.freight = freight  
        cnote.docket_charge = docket_charge 
        cnote.door_delivery_charge = door_delivery_charge  
        cnote.handling_charge = handling_charge  
        cnote.pickup_charge = pickup_charge  
        cnote.transhipment_charge = transhipment_charge  
        cnote.insurance = insurance  
        cnote.fuel_surcharge = fuel_surcharge  
        cnote.commission = commission  
        cnote.other_charge = other_charge  
        cnote.carrier_risk = carrier_risk  
        cnote.grand_total = grand_total  
        cnote.save()  
        messages.success(request, 'CNote updated successfully.')  
        return redirect('dealer:dealer_cnotes')  
    else:  
        return render(request, 'dealer/update_cnote.html', {'cnote': cnote})  

@login_required  
@user_owns_cnote  
def delete_cnote(request, cnote_id):  
    cnote = get_object_or_404(CNote, id=cnote_id)  
    if request.method == 'POST':  
        cnote.delete()  
        messages.success(request, 'CNote deleted successfully.')  
        return redirect('dealer:dealer_cnotes')  
    return render(request, 'dealer/confirm_delete.html', {'cnote': cnote})  

import logging

logger = logging.getLogger(__name__)

@login_required
@require_POST
@transaction.atomic
def create_loading_sheet(request):
    try:
        data = json.loads(request.body)
        transporter_id = data.get('transporter_id')
        cnote_ids = data.get('cnote_ids', [])

        logger.info(f"Creating loading sheet for transporter {transporter_id} with CNotes: {cnote_ids}")

        transporter = get_object_or_404(Transporter, id=transporter_id)
        dealer = request.user.dealer

        loading_sheet = LoadingSheetSummary.objects.create(
            dealer=dealer,
            transporter=transporter,
            city=dealer.city,
            total_cnote_number=len(cnote_ids),
            state=dealer.state,
            status='dispatched'
        )

        logger.info(f"Created LoadingSheetSummary with ID: {loading_sheet.ls_number}")

        total_package = 0
        total_paid_amount = 0
        total_to_pay_amount = 0

        for cnote_id in cnote_ids:
            try:
                cnote = CNotes.objects.select_for_update().get(pk=cnote_id)
                if cnote.status in ['booked', 'received_at_godown']:
                    LoadingSheetDetail.objects.create(
                        loading_sheet=loading_sheet,
                        cnote=cnote,
                        consignor_name=cnote.consignor_name,
                        consignee_name=cnote.consignee_name,
                        consignor_contact=cnote.consignor_mobile,
                        consignee_contact=cnote.consignee_mobile,
                        destination=str(cnote.delivery_destination),
                        art=cnote.total_art,
                        payment_type=cnote.payment_type,
                        amount=cnote.grand_total,
                        status='dispatched'
                    )
                    logger.info(f"Created LoadingSheetDetail for CNotes {cnote.cnote_number}")
                    
                    # Update CNotes status to 'dispatched'
                    cnote.status = 'dispatched'
                    cnote.save()
                    logger.info(f"Updated CNotes {cnote.cnote_number} status to 'dispatched'")

                    total_package += cnote.articles.count()
                    if cnote.payment_type == 'PAID':
                        total_paid_amount += cnote.grand_total
                    else:
                        total_to_pay_amount += cnote.grand_total
                else:
                    logger.warning(f"CNotes {cnote.cnote_number} status is not eligible for dispatch")
            except CNotes.DoesNotExist:
                logger.warning(f"CNotes with ID {cnote_id} does not exist")
                continue
            except Exception as e:
                logger.error(f"Error processing CNotes {cnote_id}: {str(e)}")
                continue

        loading_sheet.total_package = total_package
        loading_sheet.total_paid_amount = total_paid_amount
        loading_sheet.total_to_pay_amount = total_to_pay_amount
        loading_sheet.save()

        logger.info(f"Loading sheet {loading_sheet.ls_number} updated with totals")

        return JsonResponse({
            'message': 'Loading sheet created successfully',
            'loading_sheet_id': loading_sheet.ls_number
        })

    except Exception as e:
        logger.error(f"Error creating loading sheet: {str(e)}")
        return JsonResponse({
            'error': str(e)
        }, status=400)

@login_required
@require_POST
@transaction.atomic
def cancel_loading_sheet(request, loading_sheet_id):
    try:
        loading_sheet = get_object_or_404(LoadingSheetSummary, ls_number=loading_sheet_id, dealer=request.user.dealer)
        
        # Get all CNotes associated with this loading sheet
        cnotes = CNotes.objects.filter(loading_sheet_details__loading_sheet=loading_sheet)
        
        # Update the status of each CNote back to 'booked'
        for cnote in cnotes:
            if cnote.status == 'dispatched':
                cnote.status = 'booked'
                cnote.save()
                logger.info(f"CNotes {cnote.cnote_number} status reverted to 'booked'")
        
        # Update the loading sheet status
        loading_sheet.status = 'cancelled'
        loading_sheet.save()
        
        # Delete the loading sheet details
        LoadingSheetDetail.objects.filter(loading_sheet=loading_sheet).delete()
        
        logger.info(f"Loading sheet {loading_sheet_id} cancelled successfully")
        
        return JsonResponse({
            'success': True,
            'message': 'Loading sheet cancelled successfully.'
        })
    except Exception as e:
        logger.error(f"Error cancelling loading sheet {loading_sheet_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def update_cnote_status(request, cnote_id):
    cnote = get_object_or_404(CNotes, id=cnote_id)
    new_status = request.POST.get('status')
    if new_status:
        cnote.status = new_status
        cnote.save()
        messages.success(request, f'CNotes status updated to: {new_status}')
    else:
        messages.error(request, 'Invalid status')
    return redirect('dealer:create_cnotes')
# ... other views ...

@login_required
def mark_cnote_received(request, cnote_id):
    cnote = get_object_or_404(CNotes, id=cnote_id)
    cnote.update_status('received')  # Update status to received
    messages.success(request, 'CNotes marked as received')
    return redirect('transporter:cnote_detail', cnote_id=cnote.id)

@login_required
def create_delivery(request, cnote_id):
    cnote = get_object_or_404(CNotes, id=cnote_id)
    Delivery.objects.create(cnote=cnote)  # Create a new delivery record
    cnote.update_status('due_for_delivery')  # Update status to due for delivery
    messages.success(request, 'Delivery created and CNotes marked as due for delivery')
    return redirect('transporter:delivery_list')

@login_required
def update_delivery(request, delivery_id):
    delivery = get_object_or_404(Delivery, id=delivery_id)
    if request.method == 'POST':
        is_successful = request.POST.get('is_successful') == 'true'
        delivery.is_successful = is_successful
        delivery.delivered_at = timezone.now() if is_successful else None
        delivery.save()
        status = 'delivered' if is_successful else 'received_at_godown'
        messages.success(request , f'Delivery updated and CNotes marked as {status}')
    return redirect('transporter:delivery_list')

@login_required
def cancel_cnote(request, cnote_id):
    cnote = get_object_or_404(CNotes, id=cnote_id)
    if cnote.status != 'cancelled':
        cnote.update_status('cancelled')
        messages.success(request, 'CNotes cancelled successfully')
    else:
        messages.error(request, 'This CNotes is already cancelled')
    return redirect('dealer:cnote_detail', cnote_id=cnote.id)

@login_required
def loading_sheet(request):
    dealer = request.user.dealer  # Get the dealer associated with the logged-in user
    cnotes = CNotes.objects.filter(dealer=dealer, status__in=['booked', 'received'])
    destinations = DeliveryDestination.objects.filter(id__in=cnotes.values('delivery_destination'))

    # Fetch all transporters to pass to the template
    transporters = Transporter.objects.all()  # Fetch all transporters

    context = {
        'available_cnotes': cnotes,
        'destinations': destinations,
        'dealer': dealer,  # Ensure dealer is included in the context
        'transporters': transporters  # Pass transporters to the template
    }

    if request.method == 'POST':
        cnote_ids = request.POST.getlist('cnote_ids')  # Get the list of CNotes IDs
        transporter_id = request.POST.get('transporter_id')

        if not cnote_ids:
            messages.error(request, 'Please select at least one CNotes for the loading sheet.')
            return redirect('dealer:loading_sheet')

        if not transporter_id:
            messages.error(request, 'Please select a transporter.')
            return redirect('dealer:loading_sheet')

        try:
            transporter = Transporter.objects.get(id=transporter_id)
        except Transporter.DoesNotExist:
            messages.error(request, 'Invalid transporter selected.')
            return redirect('dealer:loading_sheet')

        # Create loading sheet logic here...

        messages.success(request, 'Loading sheet created successfully.')
        return redirect('dealer:loading_sheet_detail', loading_sheet_id=loading_sheet.id)

    return render(request, 'dealer/loading_sheet.html', context)

# Set up logging
logger = logging.getLogger(__name__)

@login_required
def booking_register_view(request):
    logger.info("üì¢ booking_register_view called")

    # Ensure the logged-in user is a dealer
    try:
        dealer = request.user.dealer
        logger.info(f"‚úÖ Dealer found: {dealer}")
    except AttributeError:
        logger.error("User is not a dealer")
        return render(request, 'dealer/booking_register.html', {'error_message': '‡§Ü‡§™ ‡§°‡•Ä‡§≤‡§∞ ‡§®‡§π‡•Ä‡§Ç ‡§π‡•à‡§Ç‡•§'})

    # Base Query for Dealer's CNotes, exclude null delivery_destination
    dealer_cnotes = CNotes.objects.filter(dealer=dealer, delivery_destination__isnull=False).order_by('-created_at')
    
    # Initialize Filters
    filters = Q()
    
    # Get Query Parameters from Request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    from_city = request.GET.get('from_city')
    to_city = request.GET.get('to_city')
    amount_min = request.GET.get('amount_min')
    amount_max = request.GET.get('amount_max')
    search_term = request.GET.get('search_term')
    cnote_number = request.GET.get('cnote_number')
    ls_number = request.GET.get('ls_number')
    payment_filter = request.GET.get('payment_type')

    # Fetch Unique To Cities for Dropdown
    to_city_list = dealer_cnotes.values_list('delivery_destination__destination_name', flat=True).distinct()
    to_city = sorted(set(city for city in to_city_list if city is not None))  # Filter out None before sorting
    to_city.insert(0, "All City")  # Add "All City" option

    # Get Selected City from Request
    selected_to_city = request.GET.get('to_city', '').strip()

    # UNIQUE payment types
    payment_types = dealer_cnotes.values_list('payment_type', flat=True).distinct()
    payment_types = sorted(set(pt for pt in payment_types if pt is not None))  # Filter out None

    # Apply Filters
    if date_from:
        try:
            filters &= Q(created_at__gte=timezone.make_aware(timezone.datetime.strptime(date_from, '%Y-%m-%d')))
        except ValueError:
            logger.warning("‚ö† Invalid date_from value")
    if date_to:
        try:
            filters &= Q(created_at__lte=timezone.make_aware(timezone.datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)))
        except ValueError:
            logger.warning("‚ö† Invalid date_to value")
    if from_city:
        filters &= Q(delivery_destination__destination_name__icontains=from_city)
    if selected_to_city and selected_to_city != "All City":
        filters &= Q(delivery_destination__destination_name=selected_to_city)
    if amount_min:
        try:
            amount_min = float(amount_min)
            filters &= Q(grand_total__gte=amount_min)
        except ValueError:
            logger.warning("‚ö† Invalid amount_min value")
    if amount_max:
        try:
            amount_max = float(amount_max)
            filters &= Q(grand_total__lte=amount_max)
        except ValueError:
            logger.warning("‚ö† Invalid amount_max value")
    if search_term:
        filters &= (Q(cnote_number__icontains=search_term) |
                    Q(consignor_name__icontains=search_term) |
                    Q(consignee_name__icontains=search_term))
    if cnote_number:
        filters &= Q(cnote_number__icontains=cnote_number)
    if ls_number:
        filters &= Q(loading_sheet_details__loading_sheet__ls_number__icontains=ls_number)
    if payment_filter and payment_filter != "All":
        filters &= Q(payment_type=payment_filter)

    # Apply Filters to Queryset
    dealer_cnotes = dealer_cnotes.filter(filters).distinct()
    logger.info(f"‚úÖ Filtered CNotes count: {dealer_cnotes.count()}")

    # Fix LS Number & DDM Number Issues
    for cnote in dealer_cnotes:
        try:
            loading_sheet_detail = cnote.loading_sheet_details.select_related('loading_sheet').first()
            cnote.loading_sheet_number = loading_sheet_detail.loading_sheet.ls_number if loading_sheet_detail else ''
        except Exception as e:
            logger.error(f"‚ö† Error accessing LS Number for CNote {cnote.cnote_number}: {str(e)}")
            cnote.loading_sheet_number = ''

    # Pagination (30 records per page)
    page = request.GET.get('page', 1)
    paginator = Paginator(dealer_cnotes, 30)
    try:
        bookings = paginator.page(page)
    except PageNotAnInteger:
        bookings = paginator.page(1)
    except EmptyPage:
        bookings = paginator.page(paginator.num_pages)

    # Calculate Summary Data
    summary_data = dealer_cnotes.exclude(grand_total__isnull=True).aggregate(
        paid_total=Sum('grand_total', filter=Q(payment_type__iregex=r'^\s*(paid)\s*$')),
        to_pay_total=Sum('grand_total', filter=Q(payment_type__iregex=r'^\s*(to\s*pay|topay)\s*$')),
        billing_total=Sum('grand_total', filter=Q(payment_type__iexact="TBB")),
        total_amount=Sum('grand_total'),
        gst_total=Sum(F('grand_total') * 0.18, output_field=FloatField()),
        grand_total=Sum(F('grand_total') * 1.18, output_field=FloatField())
    )
    for key, value in summary_data.items():
        summary_data[key] = value if value is not None else 0
    logger.info(f"‚úÖ Summary Data: {summary_data}")

    # Fetch Articles Data for Each CNote
    article_data = {}
    for cnote in dealer_cnotes:
        articles = Article.objects.filter(cnote=cnote)
        article_data[cnote.id] = {
            'total_sum': articles.aggregate(total_sum=Sum('art'))['total_sum'] or 0,
            'art_types': ' / '.join([article.art_type.art_type_name if article.art_type else 'N/A' for article in articles]) or 'N/A',
            'said_to_contain': ' / '.join([f"{article.said_to_contain or 'N/A'} ({article.art or 0})" for article in articles]) or 'N/A',
            'art_amounts': ' / '.join([str(article.art_amount or 0) for article in articles]) or '0'
        }

    # Pass Data to the Template
    return render(request, 'dealer/booking_register.html', {
        'bookings': bookings,
        'paginator': paginator,
        'to_city': to_city,
        'payment_types': payment_types,
        'summary_data': summary_data,
        'article_data': article_data,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'from_city': from_city,
            'to_city': selected_to_city,
            'amount_min': amount_min,
            'amount_max': amount_max,
            'search_term': search_term,
            'cnote_number': cnote_number,
            'ls_number': ls_number,
            'payment_type': payment_filter or 'All'
        }
    })


@login_required
def download_excel(request):
    try:
        dealer = request.user.dealer
    except AttributeError:
        messages.error(request, "‡§Ü‡§™ ‡§°‡•Ä‡§≤‡§∞ ‡§®‡§π‡•Ä‡§Ç ‡§π‡•à‡§Ç‡•§ ‡§ï‡•É‡§™‡§Ø‡§æ ‡§™‡•Å‡§®‡§É ‡§™‡•ç‡§∞‡§Ø‡§æ‡§∏ ‡§ï‡§∞‡•á‡§Ç‡•§")
        return redirect('dealer:booking_register')

    # ‚úÖ **Same Filters Apply ‡§ï‡§∞‡•á‡§Ç ‡§ú‡•ã `booking_register_view` ‡§Æ‡•á‡§Ç ‡§π‡•à‡§Ç**
    filters = Q(dealer=dealer)

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    from_city = request.GET.get('from_city')
    to_city = request.GET.get('to_city')
    amount_min = request.GET.get('amount_min')
    amount_max = request.GET.get('amount_max')
    search_term = request.GET.get('search_term')
    cnote_number = request.GET.get('cnote_number')
    ls_number = request.GET.get('ls_number')
    payment_filter = request.GET.get('payment_type')

    if date_from:
        filters &= Q(created_at__gte=timezone.make_aware(timezone.datetime.strptime(date_from, '%Y-%m-%d')))
    if date_to:
        filters &= Q(created_at__lte=timezone.make_aware(timezone.datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)))

    if from_city:
        filters &= Q(delivery_destination__destination_name__icontains=from_city)

    if to_city and to_city != "All City":
        filters &= Q(delivery_destination__destination_name=to_city)

    if amount_min:
        try:
            amount_min = float(amount_min)
            filters &= Q(grand_total__gte=amount_min)
        except ValueError:
            pass

    if amount_max:
        try:
            amount_max = float(amount_max)
            filters &= Q(grand_total__lte=amount_max)
        except ValueError:
            pass

    if search_term:
        filters &= (Q(cnote_number__icontains=search_term) |
                    Q(consignor_name__icontains=search_term) |
                    Q(consignee_name__icontains=search_term))

    if cnote_number:
        filters &= Q(cnote_number__icontains=cnote_number)

    if ls_number:
        filters &= Q(loading_sheet_details__loading_sheet__ls_number__icontains=ls_number)

    if payment_filter and payment_filter != "All":
        filters &= Q(payment_type=payment_filter)

    # ‚úÖ **Filtered CNotes Fetch ‡§ï‡§∞‡•á‡§Ç**
    dealer_cnotes = CNotes.objects.filter(filters).order_by('-created_at')

    # ‚úÖ **Article Data Fetch ‡§ï‡§∞‡•á‡§Ç**
    article_data = {}
    for cnote in dealer_cnotes:
        articles = Article.objects.filter(cnote=cnote)
        article_data[cnote.id] = {
            "total_art": len(articles),
            "art_types": "/".join([article.art_type.art_type_name for article in articles if article.art_type]),
            "said_to_contain": "/".join([article.said_to_contain for article in articles if article.said_to_contain]),
            "art_amounts": "/".join([str(article.art_amount) for article in articles if article.art_amount]),
        }

    # ‚úÖ **Excel File Create ‡§ï‡§∞‡•á‡§Ç**
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("Booking Register")

    # ‚úÖ **Excel Headers Define ‡§ï‡§∞‡•á‡§Ç (Updated)**
    headers = [
        'Sr. No', 'CNote Number', 'Booking Type', 'Delivery Destination', 'Consignor Name', 'Consignee Name',
        'Payment Type', 'Grand Total', 'Created Date', 'Created Time',  # ‚úÖ Yeh do naye columns
        'Eway Bill Number', 'Actual Weight', 'Charged Weight', 'Weight Rate', 'Weight Amount', 'Fix Amount',
        'Invoice Number', 'Declared Value', 'Risk Type', 'POD Required', 'Freight', 'Docket Charge',
        'Door Delivery Charge', 'Handling Charge', 'Pickup Charge', 'Transhipment Charge',
        'Insurance', 'Fuel Surcharge', 'Commission', 'Other Charge', 'Carrier Risk',
        'Delivery Type', 'Delivery Method', 'Status', 'Total Article', 'Art Types',
        'Said to Contain', 'Art Amounts', 'Consignor GST', 'Consignee GST', 'LS Number'
    ]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header)  # ‚úÖ Headers ‡§ï‡•ã First Row ‡§Æ‡•á‡§Ç Add ‡§ï‡§∞‡•á‡§Ç

    # ‚úÖ **Table Data ‡§ï‡•ã Excel ‡§Æ‡•á‡§Ç ‡§≤‡§ø‡§ñ‡•á‡§Ç**
    for row, cnote in enumerate(dealer_cnotes, start=1):
        article_info = article_data.get(cnote.id, {"total_art": 0, "art_types": "N/A", "said_to_contain": "N/A", "art_amounts": "0"})

        # ‚úÖ "Created At" ka Date aur Time alag-alag nikalna
        created_date = cnote.created_at.strftime('%d-%m-%Y') if cnote.created_at else 'N/A'  # ‚úÖ dd-mm-yyyy format
        created_time = cnote.created_at.strftime('%H:%M:%S') if cnote.created_at else 'N/A'  # ‚úÖ Same time format

        data = [
            row, cnote.cnote_number, cnote.booking_type,
            cnote.delivery_destination.destination_name if cnote.delivery_destination else 'N/A',
            cnote.consignor_name, cnote.consignee_name, cnote.payment_type, cnote.grand_total,
            created_date,  # ‚úÖ Yeh "Created Date" ke liye hai
            created_time,  # ‚úÖ Yeh "Created Time" ke liye hai
            cnote.eway_bill_number,
            cnote.actual_weight, cnote.charged_weight, cnote.weight_rate,
            cnote.weight_amount, cnote.fix_amount, cnote.invoice_number, cnote.declared_value,
            cnote.risk_type, cnote.pod_required, cnote.freight, cnote.docket_charge,
            cnote.door_delivery_charge, cnote.handling_charge, cnote.pickup_charge,
            cnote.transhipment_charge, cnote.insurance, cnote.fuel_surcharge, cnote.commission,
            cnote.other_charge, cnote.carrier_risk, cnote.delivery_type, cnote.delivery_method,
            cnote.status,
            article_info["total_art"],  # ‚úÖ Total Articles
            article_info["art_types"],  # ‚úÖ Art Types
            article_info["said_to_contain"],  # ‚úÖ Said to Contain
            article_info["art_amounts"],  # ‚úÖ Art Amounts
            cnote.consignor_gst, cnote.consignee_gst,
            cnote.loading_sheet_details.first().loading_sheet.ls_number if cnote.loading_sheet_details.exists() else 'N/A'
        ]

        for col, value in enumerate(data):
            worksheet.write(row, col, value)

    workbook.close()
    output.seek(0)

    # ‚úÖ **Excel File ‡§ï‡•ã Response ‡§Æ‡•á‡§Ç Return ‡§ï‡§∞‡•á‡§Ç**
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=booking_register.xlsx'
    return response

@login_required
def cnote_options(request, cnote_id):
    cnote = get_object_or_404(CNotes, id=cnote_id)
    if request.method == 'POST':
        option = request.POST.get('option')
        if option == 'pdf':
            pdf_file = render_to_pdf('dealer/cnote_pdf.html', {'cnote': cnote})
            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="cnote_{cnote_id}.pdf"'
            return response
        elif option == 'print':
            return redirect('create_cnotes')  # Adjust as needed
    return render(request, 'dealer/cnote_options.html', {'cnote': cnote})


@login_required
def view_pending_pickups(request, transporter_id):
    # Get the transporter object based on the ID
    transporter = get_object_or_404(Transporter, id=transporter_id)

    # Fetch the pending pickups for this transporter
    pending_pickups = Pickup.objects.filter(transporter=transporter, status='pending')

    # Render the template with the pending pickups
    return render(request, 'dealer/view_pending_pickups.html', {
        'transporter': transporter,
        'pending_pickups': pending_pickups
    })



@login_required
def view_pending_pickups(request, transporter_id):
    # Get the transporter object based on the ID
    transporter = get_object_or_404(Transporter, id=transporter_id)

    # Fetch the pending pickups for this transporter
    pending_pickups = Pickup.objects.filter(transporter=transporter, status='pending')

    # Render the template with the pending pickups
    return render(request, 'dealer/view_pending_pickups.html', {
        'transporter': transporter,
        'pending_pickups': pending_pickups
    })




@login_required
def view_pickups(request):
    pickups = Pickup.objects.all()  # Get all pickups
    return render(request, 'dealer/view_pickups.html', {'pickups': pickups})




def mark_pickup(request, cnote_id):
    # Logic to mark the pickup
    pickup = get_object_or_404(Pickup, cnote_id=cnote_id)
    # Update the pickup status or any other logic you need
    pickup.status = 'picked_up'  # Example status update
    pickup.save()
    return render(request, 'dealer/pickup_marked.html', {'pickup': pickup})

def add_dealer(request):
    # ‡§Ø‡§π‡§æ‡§Å ‡§™‡§∞ ‡§°‡•Ä‡§≤‡§∞ ‡§ú‡•ã‡§°‡§º‡§®‡•á ‡§ï‡•Ä ‡§≤‡•â‡§ú‡§ø‡§ï ‡§°‡§æ‡§≤‡•á‡§Ç
    return render(request, 'dealer/add_dealer.html')  # ‡§°‡•Ä‡§≤‡§∞ ‡§ú‡•ã‡§°‡§º‡§®‡•á ‡§ï‡§æ ‡§ü‡•á‡§Æ‡•ç‡§™‡§≤‡•á‡§ü


def print_pdf(request, cnote_id):
    # ‡§Ø‡§π‡§æ‡§Å ‡§™‡§∞ PDF ‡§™‡•ç‡§∞‡§ø‡§Ç‡§ü ‡§ï‡§∞‡§®‡•á ‡§ï‡•Ä ‡§≤‡•â‡§ú‡§ø‡§ï ‡§°‡§æ‡§≤‡•á‡§Ç
    return render(request, 'dealer/print_pdf.html', {'cnote_id': cnote_id})  # PDF ‡§™‡•ç‡§∞‡§ø‡§Ç‡§ü ‡§ï‡§∞‡§®‡•á ‡§ï‡§æ ‡§ü‡•á‡§Æ‡•ç‡§™‡§≤‡•á‡§ü


def download_pdf(request, cnote_id):
    # ‡§Ø‡§π‡§æ‡§Å ‡§™‡§∞ PDF ‡§°‡§æ‡§â‡§®‡§≤‡•ã‡§° ‡§ï‡§∞‡§®‡•á ‡§ï‡•Ä ‡§≤‡•â‡§ú‡§ø‡§ï ‡§°‡§æ‡§≤‡•á‡§Ç
    cnote = CNotes.objects.get(id=cnote_id)  # ‡§â‡§¶‡§æ‡§π‡§∞‡§£ ‡§ï‡•á ‡§≤‡§ø‡§è, CNotes ‡§Æ‡•â‡§°‡§≤ ‡§∏‡•á ‡§°‡•á‡§ü‡§æ ‡§™‡•ç‡§∞‡§æ‡§™‡•ç‡§§ ‡§ï‡§∞‡§®‡§æ
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cnote_{cnote_id}.pdf"'
    
    # ‡§Ø‡§π‡§æ‡§Å ‡§™‡§∞ PDF ‡§ú‡§®‡§∞‡•á‡§ü ‡§ï‡§∞‡§®‡•á ‡§ï‡•Ä ‡§≤‡•â‡§ú‡§ø‡§ï ‡§°‡§æ‡§≤‡•á‡§Ç
    # ‡§â‡§¶‡§æ‡§π‡§∞‡§£ ‡§ï‡•á ‡§≤‡§ø‡§è, ‡§Ü‡§™ ReportLab ‡§Ø‡§æ ‡§ï‡§ø‡§∏‡•Ä ‡§Ö‡§®‡•ç‡§Ø ‡§≤‡§æ‡§á‡§¨‡•ç‡§∞‡•á‡§∞‡•Ä ‡§ï‡§æ ‡§â‡§™‡§Ø‡•ã‡§ó ‡§ï‡§∞ ‡§∏‡§ï‡§§‡•á ‡§π‡•à‡§Ç

    return response  # PDF ‡§´‡§º‡§æ‡§á‡§≤ ‡§ï‡•ã ‡§µ‡§æ‡§™‡§∏ ‡§ï‡§∞‡•á‡§Ç


@login_required
def get_cnote_details(request, cnote_number):
    cnote = get_object_or_404(CNotes, cnote_number=cnote_number)
    data = {
        'cnote_number': cnote.cnote_number,
        'created_at': cnote.created_at.isoformat(),
        'payment_type': cnote.payment_type,
        'booking_type': cnote.booking_type,
        'delivery_type': cnote.delivery_type,
        'delivery_method': cnote.delivery_method,
        'delivery_destination': cnote.delivery_destination,
        'eway_bill_number': cnote.eway_bill_number,
        'consignor_name': cnote.consignor_name,
        'consignor_mobile': cnote.consignor_mobile,
        'consignor_gst': cnote.consignor_gst,
        'consignor_address': cnote.consignor_address,
        'consignee_name': cnote.consignee_name,
        'consignee_mobile': cnote.consignee_mobile,
        'consignee_gst': cnote.consignee_gst,
        'consignee_address': cnote.consignee_address,
        'articles': [
            {
                'name': article.name,
                'quantity': article.quantity,
                'type': article.type,
                'said_to_contain': article.said_to_contain,
                'amount': float(article.amount)
            } for article in cnote.articles.all()
        ],
        'freight': float(cnote.freight),
        'docket_charge': float(cnote.docket_charge),
        'door_delivery_charge': float(cnote.door_delivery_charge),
        'handling_charge': float(cnote.handling_charge),
        'other_charge': float(cnote.other_charge),
        'grand_total': float(cnote.grand_total),
        'dealer': {
            'logo_url': cnote.dealer.logo.url if cnote.dealer.logo else None
        }
    }
    return JsonResponse(data)

@login_required
def cnote_success(request, cnote_number):
    try:
        # CNotes ‡§ë‡§¨‡•ç‡§ú‡•á‡§ï‡•ç‡§ü ‡§ï‡•ã ‡§∏‡§Ç‡§¨‡§Ç‡§ß‡§ø‡§§ ‡§Ü‡§∞‡•ç‡§ü‡§ø‡§ï‡§≤‡•ç‡§∏ ‡§î‡§∞ ‡§°‡•Ä‡§≤‡§∞ ‡§ï‡•á ‡§∏‡§æ‡§• ‡§≤‡§æ‡§®‡§æ
        cnote = CNotes.objects.prefetch_related('articles').select_related('dealer').get(cnote_number=cnote_number)
        
        # ‡§∏‡§≠‡•Ä ‡§∏‡§Ç‡§¨‡§Ç‡§ß‡§ø‡§§ ‡§Ü‡§∞‡•ç‡§ü‡§ø‡§ï‡§≤‡•ç‡§∏ ‡§ï‡•ã ‡§≤‡§æ‡§®‡§æ
        articles = cnote.articles.all()  # ‡§Ø‡§π ‡§∏‡§≠‡•Ä ‡§Ü‡§∞‡•ç‡§ü‡§ø‡§ï‡§≤‡•ç‡§∏ ‡§ï‡•ã ‡§≤‡§æ‡§è‡§ó‡§æ
        
        # Total Art ‡§ï‡•Ä ‡§ó‡§£‡§®‡§æ ‡§ï‡§∞‡§®‡§æ
        total_art = sum(article.art for article in articles if article.art is not None)  # None ‡§ö‡•á‡§ï ‡§ï‡§∞‡•á‡§Ç
        
        # ‡§ü‡•á‡§Æ‡•ç‡§™‡§≤‡•á‡§ü ‡§Æ‡•á‡§Ç cnote, dealer, articles ‡§î‡§∞ total_art ‡§™‡§æ‡§∏ ‡§ï‡§∞‡§®‡§æ
        return render(request, 'dealer/cnote_success.html', {
            'cnote': cnote,
            'dealer': cnote.dealer,
            'articles': articles,  # articles ‡§ï‡•ã ‡§™‡§æ‡§∏ ‡§ï‡§∞‡§®‡§æ
            'total_art': total_art,  # total_art ‡§ï‡•ã ‡§™‡§æ‡§∏ ‡§ï‡§∞‡§®‡§æ
        })
    except CNotes.DoesNotExist:
        messages.error(request, 'CNote not found.')
        return redirect('dealer:create_cnotes')


logger = logging.getLogger(__name__)

@login_required
def view_cnote(request, cnote_number):
    try:
        # Fetch the CNote
        cnote = get_object_or_404(CNotes, cnote_number=cnote_number)
        
        # Fetch the dealer information from the CNote
        dealer = cnote.dealer
        
        # Fetch related articles
        #articles = cnote.articles.all()
        articles = Article.objects.filter(cnote=cnote)  # Fetch related articles


        # Fetch all destinations for the dropdown
        destinations = DeliveryDestination.objects.all()

        # Pass all required data to the template
        return render(request, 'dealer/view_cnote.html', {
            'cnote': cnote,
            'dealer': dealer,
            'articles': articles,
            'destinations': destinations,
            #'article_types': article_types, # type: ignore
        })

    except CNotes.DoesNotExist:
        logger.error(f"CNote does not exist: {cnote_number}")
        messages.error(request, 'CNote does not exist.')
        return redirect('dealer:create_cnotes')
    
    except Exception as e:
        logger.error(f"Error in view_cnote: {str(e)}")
        messages.error(request, 'An error occurred while fetching the CNote details.')
        return redirect('dealer:dashboard')

# dealer_app/views.py

logger = logging.getLogger(__name__)

@login_required
def fetch_pending_cnotes(request):
    dealer = request.user.dealer
    cnotes = CNotes.objects.filter(
        dealer=dealer, 
        status__in=['booked', 'received_at_godown']
    ).values(
        'id', 'created_at', 'cnote_number', 'delivery_destination__destination_name',
        'consignor_name', 'consignee_name', 'total_art', 'actual_weight', 
        'charged_weight', 'booking_type', 'freight', 'grand_total',
        'payment_type'
    ).annotate(
        total_articles=Sum('articles__art')  # ‚úÖ Change 'total_art' to 'total_articles'
    )
    
    # Convert queryset to list and ensure total_art is not None
    cnotes_list = list(cnotes)
    for cnote in cnotes_list:
        if cnote['total_art'] is None:
            cnote['total_art'] = 0
            
    return JsonResponse(cnotes_list, safe=False)

logger = logging.getLogger(__name__)
@csrf_exempt
def fetch_transporters(request):
    transporters = Transporter.objects.all()
    
    # Create XML structure
    root = Element('transporters')
    for transporter in transporters:
        transporter_elem = SubElement(root, 'transporter')
        transporter_elem.set('id', str(transporter.id))
        name_elem = SubElement(transporter_elem, 'name')
        name_elem.text = transporter.name

    # Convert XML to string
    xml_str = tostring(root, encoding='unicode')
    
    # Return XML response
    return HttpResponse(xml_str, content_type='application/xml')

@login_required
def home(request):
    # Logic specific to the dealer dashboard
    return render(request, 'transporter_app/home.html')  # Adjust the template path as needed


def get_cities(request):
    if request.is_ajax():
        cities = City.objects.values_list('name', flat=True)
        return JsonResponse(list(cities), safe=False)
    return JsonResponse([], safe=False)

def fetch_cities(request):
    """
    Fetches all city names from the City model and returns them as a JSON response.
    """
    if request.method == "GET":
        try:
            # Fetch city names from the City model
            cities = City.objects.values_list('name', flat=True).order_by('name')  # Order by name for better usability
            city_list = list(cities)  # Convert QuerySet to a list
            return JsonResponse({'success': True, 'cities': city_list}, safe=False)
        except Exception as e:
            # Handle any unexpected errors
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    else:
        # Return a 405 error for non-GET requests
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

logger = logging.getLogger(__name__)

@login_required
@csrf_exempt
def send_pickup_request(request):
    if request.method == 'POST':
        logger.debug("Received POST request for sending pickup request.")
        logger.debug(f"POST data: {request.POST}")
        
        transporter_name = request.POST.get('transporter_name')
        cnote_ids = request.POST.get('cnote_ids', '')

        logger.debug(f"Transporter name: {transporter_name}, CNotes IDs: {cnote_ids}")

        if not transporter_name:
            messages.error(request, 'No transporter name provided.')
            logger.warning("No transporter name provided.")
            return redirect('dealer:loading_sheet')

        if isinstance(cnote_ids, str):
            cnote_ids = [id.strip() for id in cnote_ids.split(',') if id.strip()]
        
        if not cnote_ids:
            messages.error(request, 'No valid CNotes IDs provided.')
            logger.warning("No valid CNotes IDs provided.")
            return redirect('dealer:loading_sheet')

        try:
            transporter = Transporter.objects.get(name=transporter_name)
        except Transporter.DoesNotExist:
            messages.error(request, f'Transporter with name "{transporter_name}" does not exist.')
            logger.warning(f"Transporter with name '{transporter_name}' does not exist.")
            return redirect('dealer:loading_sheet')

        dealer = request.user.dealer
        
        try:
            loading_sheet_summary = LoadingSheetSummary.objects.create(
                dealer=dealer,
                transporter=transporter,
                total_cnote=len(cnote_ids),
                status='dispatched'
            )

            valid_cnote_ids = []
            total_art = 0
            total_paid_amount = total_topay_amount = total_tbb_amount = total_foc_amount = 0

            for cnote_id in cnote_ids:
                try:
                    cnote = CNotes.objects.get(id=int(cnote_id))
                    if cnote.status in ['booked', 'received_at_godown']:  # Update: Added status check
                        valid_cnote_ids.append(cnote_id)

                        LoadingSheetDetail.objects.create(
                            loading_sheet=loading_sheet_summary,
                            cnote=cnote,
                            consignor_name=cnote.consignor_name,
                            consignee_name=cnote.consignee_name,
                            consignor_contact=cnote.consignor_mobile,
                            consignee_contact=cnote.consignee_mobile,
                            destination=str(cnote.delivery_destination),
                            art=cnote.total_art,
                            payment_type=cnote.payment_type,
                            amount=cnote.grand_total,
                            status='dispatched'
                        )

                    cnote.status = 'dispatched'
                    cnote.save()
                    logger.info(f"CNotes {cnote_id} status updated to 'dispatched'")

                    total_art += cnote.total_art
                    if cnote.payment_type == 'PAID':
                        total_paid_amount += cnote.grand_total
                    elif cnote.payment_type == 'TO PAY':
                        total_topay_amount += cnote.grand_total
                    elif cnote.payment_type == 'TBB':
                        total_tbb_amount += cnote.grand_total
                    elif cnote.payment_type == 'FOC':
                        total_foc_amount += cnote.grand_total

                    logger.debug(f"LoadingSheetDetail created for CNotes ID: {cnote_id}")
                except ValueError:
                    logger.warning(f"Invalid CNotes ID format: {cnote_id}")
                except CNotes.DoesNotExist:
                    logger.warning(f"CNotes with ID {cnote_id} does not exist")
                except Exception as e:
                    logger.error(f"Error processing CNotes {cnote_id}: {str(e)}")

            loading_sheet_summary.total_art = total_art
            loading_sheet_summary.total_paid_amount = total_paid_amount
            loading_sheet_summary.total_topay_amount = total_topay_amount
            loading_sheet_summary.total_tbb_amount = total_tbb_amount
            loading_sheet_summary.total_foc_amount = total_foc_amount
            loading_sheet_summary.save()

            logger.info(f"Loading sheet {loading_sheet_summary.ls_number} created successfully")

            messages.success(request, f'Loading sheet {loading_sheet_summary.ls_number} created successfully!')

            # Redirect to the mf_print.html page using ls_number
            return redirect('dealer:mf_print', loading_sheet_number=loading_sheet_summary.ls_number)
        
        except Exception as e:
            logger.error(f"Error creating loading sheet: { str(e)}")
            messages.error(request, f'Error creating loading sheet: {str(e)}')
            return redirect('dealer:loading_sheet')

    messages.error(request, 'Invalid request method.')
    return redirect('dealer:loading_sheet')

@login_required
def accept_loading_sheet(request, sheet_id):
    loading_sheet = get_object_or_404(LoadingSheet, id=sheet_id)
    for cnote in loading_sheet.cnotes.all():
        cnote.status = 'received'
        cnote.save()
    messages.success(request, 'CNotes acknowledged by transporter.')
    return redirect('transporter:home')

@login_required
def mark_as_delivered(request, cnote_id):
    cnote = get_object_or_404(CNotes, id=cnote_id)
    Delivery.objects.create(cnote=cnote, is_successful=True)
    messages.success(request, 'CNotes marked as delivered.')
    return redirect('transporter:delivery_list')

@login_required
def transporter_list(request):
    # Fetch all transporters
    transporters = Transporter.objects.all()

    # Render the template with the correct path
    return render(request, 'transporter/transporter_list.html', {'transporters': transporters})

@login_required
def search(request):
    dealer = request.user.dealer  # Fetch the logged-in dealer's information
    return render(request, 'dealer/search.html', {'dealer': dealer})  # Pass dealer info to the template



@login_required
def search_loading_sheets(request):
        # Get the logged-in dealer
    try:
        dealer = request.user.dealer
    except AttributeError:
        # If the user is not a dealer, show an error or redirect
        return redirect('login')

    if request.method == 'POST':
        search_type = request.POST.get('type')
        query = request.POST.get('query', '').strip()
        
        if search_type == 'ls':
            dealer = request.user.dealer
            if query:
                # Search for a specific loading sheet
                loading_sheets = LoadingSheetSummary.objects.filter(
                    dealer=dealer,
                    ls_number__icontains=query
                )
            else:
                # Show all loading sheets for the dealer
                loading_sheets = LoadingSheetSummary.objects.filter(dealer=dealer)
            
            results = loading_sheets.order_by('-created_at')
        else:
            results = []
    else:
        results = []
    
    return render(request, 'dealer/search.html', {
        'results': results,
        'query': query if 'query' in locals() else '',
    })

logger = logging.getLogger(__name__)

@login_required
def mf_print(request, loading_sheet_number):
    try:
        # ‚úÖ Get Loading Sheet
        loading_sheet = get_object_or_404(LoadingSheetSummary, ls_number=loading_sheet_number)

        # ‚úÖ Fetch Loading Sheet Details with Total Art Calculation
        

        loading_sheet_details = (
            LoadingSheetDetail.objects.filter(loading_sheet=loading_sheet)
            .select_related("cnote")
            .annotate(
                total_art=Sum("cnote__articles__art"),  
                invoice_number=StringAgg("cnote__invoice_number", delimiter=", ")  # ‚úÖ Fixed
            )
        )


        # ‚úÖ Calculate Summary Data - Direct approach
        summary = {
            "Paid": {"count": 0, "pkgs": 0, "amount": 0},
            "To_Pay": {"count": 0, "pkgs": 0, "amount": 0},
            "TBB": {"count": 0, "pkgs": 0, "amount": 0},
        }
        
        # ‚úÖ Process each detail record individually to ensure accurate counting
        for detail in loading_sheet_details:
            payment_type = detail.payment_type.upper().replace(" ", "_")
            
            # Map payment types to our standard keys
            if payment_type in ["PAID"]:
                key = "Paid"
            elif payment_type in ["TOPAY", "TO_PAY"]:
                key = "To_Pay"
            elif payment_type in ["TBB"]:
                key = "TBB"
            else:
                continue
                
            # ‚úÖ Increment counts and sums
            summary[key]["count"] += 1
            
            # ‚úÖ Use the total_art directly from the detail record
            # This ensures we get the exact art count from the CNotes
            summary[key]["pkgs"] += detail.total_art or 0
            summary[key]["amount"] += detail.amount or 0

        # ‚úÖ Pass Context to Template
        context = {
            "loading_sheet": loading_sheet,
            "loading_sheet_details": loading_sheet_details,
            "dealer": loading_sheet.dealer,
            "transporter": loading_sheet.transporter,
            "current_date": timezone.now(),
            "summary": summary,  # ‚úÖ Clean dictionary passed
        }

        return render(request, "dealer/mf_print.html", context)

    except Exception as e:
        logger.error(f"Error rendering MF Print page: {str(e)}")
        messages.error(request, "Error loading the print page. Please try again.")
        return redirect("dealer:loading_sheet")

    
@login_required
@require_POST
def cancel_loading_sheet(request, loading_sheet_id):
    try:
        loading_sheet = get_object_or_404(LoadingSheetSummary, ls_number=loading_sheet_id, dealer=request.user.dealer)
        
        # Get all CNotes associated with this loading sheet
        cnotes = CNotes.objects.filter(loading_sheet_details__loading_sheet=loading_sheet)
        
        # Update the status of each CNote
        for cnote in cnotes:
            if cnote.status == 'dispatched':
                # Revert the status to 'booked' or the appropriate previous state
                cnote.status = 'booked'
                cnote.save()
        
        # Update the loading sheet status
        loading_sheet.status = 'cancelled'
        loading_sheet.save()
        
        # Delete the loading sheet details
        LoadingSheetDetail.objects.filter(loading_sheet=loading_sheet).delete()
        
        return JsonResponse({'success': True, 'message': 'Loading sheet cancelled successfully.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})



logger = logging.getLogger(__name__)

@login_required
def create_cnotes(request):
    # Get the logged-in dealer
    try:
        dealer = request.user.dealer  # Assuming you have a OneToOneField relationship
        logger.info(f"Dealer: {dealer}")
    except Dealer.DoesNotExist:
        logger.error('Dealer not found.')
        return JsonResponse({'success': False, 'error': 'Dealer not found.'})

    # Fetch the last CNote created by the dealer
    last_cnote = CNotes.objects.filter(dealer=dealer).order_by('-created_at').first()

     # **‚¨áÔ∏è Fetch selected destination & checkbox value**
    delivery_destination = request.GET.get("delivery_destination", "").strip().upper()
    destination_wise = request.GET.get("destination_wise", "true") == "true"  # Default True
    consignee_list = []

    if delivery_destination and destination_wise:
        consignee_list = PartyMaster.objects.filter(
            Q(city__iexact=delivery_destination) |  
            Q(address__icontains=delivery_destination) |  
            Q(remark__icontains=delivery_destination)  
        ).order_by("party_name")
    else:
        consignee_list = PartyMaster.objects.all().order_by("party_name")  # ‚úÖ All parties if unchecked

    if request.method == 'POST':
        logger.info("POST request received.")
        logger.info(f"POST data: {request.POST}")

        try:   
            # Process charge fields with explicit mapping
            charge_fields = {
                'freight': 'freight',
                'docket_charge': 'docket_charge',
                'door_delivery_charge': 'door_delivery_charge',
                'handling_charge': 'handling_charge',
                'pickup_charge': 'pickup_charge',
                'transhipment_charge': 'transhipment_charge',
                'insurance': 'insurance',
                'fuel_surcharge': 'fuel_surcharge',
                'commission': 'commission',
                'other_charge': 'other_charge',
                'carrier_risk': 'carrier_risk',
                'grand_total': 'grand_total'
            }
            
            charges = {}
            for field, db_field in charge_fields.items():
                try:
                    value = request.POST.get(field, '0')
                    charges[db_field] = float(value)
                    logger.info(f"Processed {field}: {charges[db_field]}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error converting {field}: {str(e)}")
                    charges[db_field] = 0.0  # Default to 0.0 if conversion fails

            # Create CNotes object with all fields
            cnote = CNotes(
                dealer=dealer,
                created_at=timezone.now(),
                booking_type=request.POST.get('booking_type'),
                delivery_type=request.POST.get('delivery_type'),
                delivery_method=request.POST.get('delivery_method'),
                delivery_destination=DeliveryDestination.objects.get_or_create(
                    destination_name=request.POST.get('delivery_destination')
                )[0],
                eway_bill_number=request.POST.get('ewayBillNumber'),
                consignor_name=request.POST.get('consignor_name'),
                consignor_mobile=request.POST.get('consignor_mobile'),
                consignor_gst=request.POST.get('consignorGST'),
                consignor_address=request.POST.get('consignorAddress'),
                consignee_name=request.POST.get('consignee_name'),
                consignee_mobile=request.POST.get('consignee_mobile'),
                consignee_gst=request.POST.get('consigneeGST'),
                consignee_address=request.POST.get('consigneeAddress'),
                actual_weight=float(request.POST.get('actual_weight', 0)),
                charged_weight=float(request.POST.get('charged_weight', 0)),
                weight_rate=float(request.POST.get('weight_rate', 0)),
                weight_amount=float(request.POST.get('weight_amount', 0)),
                fix_amount=float(request.POST.get('fix_amount', 0)),
                invoice_number=request.POST.get('invoice_number'),
                declared_value=float(request.POST.get('declared_value', 0)),
                risk_type=request.POST.get('risk_type'),
                pod_required=request.POST.get('pod_required'),
                payment_type=request.POST.get('paymentType'),
                **charges  # Unpack all charge fields
            )
            
            # Save the CNotes object
            cnote.save()
            logger.info(f"CNotes saved successfully. ID: {cnote.id}")

            # Process article details from the new format
            articles_data = []
            index = 0
            while True:
                article_type = request.POST.get(f'articles[{index}][article]')
                if article_type is None:
                    break
                
                art = request.POST.get(f'articles[{index}][art]')
                art_type = request.POST.get(f'articles[{index}][artType]')
                said_to_contain = request.POST.get(f'articles[{index}][saidToContain]')
                total_amt = request.POST.get(f'articles[{index}][totalAmt]', '0')
                
                articles_data.append({
                    'article_type': article_type,
                    'art': art,
                    'art_type': art_type,
                    'said_to_contain': said_to_contain,
                    'art_amount': float(total_amt)  # Using total_amt as art_amount
                })
                index += 1

            # Create Article objects
            for article_data in articles_data:
                Article.objects.create(
                    cnote=cnote,
                    article_type=article_data['article_type'],
                    art=article_data['art'],
                    art_type=ArtType.objects.get_or_create(
                        art_type_name=article_data['art_type']
                    )[0],
                    said_to_contain=article_data['said_to_contain'],
                    art_amount=article_data['art_amount']
                )
                logger.info(f"Created article: {article_data}")

            # Return success response with redirect URL
            return JsonResponse({
                'success': True,
                'message': f'CNotes created successfully with number: {cnote.cnote_number}',
                'redirect_url': reverse('dealer:cnote_success', kwargs={'cnote_number': cnote.cnote_number})
            })

        except Exception as e:
            logger.error(f"Error saving CNotes: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': f'Error creating CNotes: {str(e)}'
            })

    # GET request - render the form
    context = {
        'dealer': dealer,
        'destinations': DeliveryDestination.objects.all(),
        'cities': City.objects.all(),
        'last_cnote': last_cnote,  # Pass the last CNote to the template
    }
    return render(request, 'dealer/create_cnotes.html', context)


# ‚úÖ STEP 3: Create `print_with_qr` View for Dual QR Logic



import base64
from io import BytesIO
import qrcode
from django.shortcuts import get_object_or_404, render
from .models import CNotes, Parcel, Article
# ‚úÖ STEP 3: Create `print_with_qr` View for Dual QR Logic
import base64
from io import BytesIO
import qrcode
from django.shortcuts import get_object_or_404, render
from .models import CNotes, Parcel, Article, QRPrinterSetting
from PIL import Image, ImageWin
import platform
if platform.system() == "Windows":
    import win32print
else:
    win32print = None
import win32ui
from django.contrib.auth.decorators import login_required
import platform
if platform.system() == "Windows":
    import win32print
else:
    win32print = None
from .models import QRPrinterSetting
from .forms import QRPrinterSelectionForm
from .utils import generate_qr_base64, send_qr_to_printer_using_html
from django.template.loader import render_to_string  # ‚úÖ ‡§Ø‡§π main import ‡§π‡•à



import imgkit
import tempfile
import os

def send_qr_to_printer_using_html(html_string, printer_name):
    try:
        print("\nüõ†Ô∏è [STEP 1] Saving HTML content to file...")

        # Step 1: Save HTML to file
        with tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8') as html_file:
            html_file.write(html_string)
            html_path = html_file.name

        print(f"‚úÖ HTML saved at: {html_path}")

        # Step 2: Output image path
        img_path = html_path.replace(".html", ".jpg")
        print(f"üì∏ Image will be saved at: {img_path}")

        # Step 3: Convert HTML to image using wkhtmltoimage
        options = {
            'format': 'jpg',
            'zoom': '1.3',
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'quality': '100',
        }

        config = imgkit.config()  # uses system-installed wkhtmltoimage
        local_file_url = f"file:///{html_path.replace(os.sep, '/')}"  # file:// works for Windows
        print(f"üîó Loading URL: {local_file_url}")
        imgkit.from_url(local_file_url, img_path, options=options, config=config)

        print("‚úÖ Image generated successfully.")

        # Step 4: Load image and send to printer
        image = Image.open(img_path)
        hprinter = win32print.OpenPrinter(printer_name)
        hdc = win32ui.CreateDC()
        hdc.CreatePrinterDC(printer_name)
        hdc.StartDoc("Parcel QR Label")
        hdc.StartPage()

        dib = ImageWin.Dib(image)
        dib.draw(hdc.GetHandleOutput(), (0, 0, image.width, image.height))

        hdc.EndPage()
        hdc.EndDoc()
        hdc.DeleteDC()
        win32print.ClosePrinter(hprinter)

        print(f"üñ®Ô∏è Sent image to printer: {printer_name}")

        # Step 5: Clean up
        os.remove(html_path)
        os.remove(img_path)

        return True

    except Exception as e:
        print(f"‚ùå QR HTML Print Error: {e}")
        import traceback
        traceback.print_exc()
        return False

@login_required
def select_qr_printer(request):
    dealer = request.user.dealer

    # List installed printers
    installed = [printer[2] for printer in win32print.EnumPrinters(2)]
    installed.insert(0, "‚ùå No QR Printer (Use A4 Sheet)")

    # Get saved setting
    setting = QRPrinterSetting.objects.filter(dealer=dealer).first()
    saved_printer = setting.printer_name if setting else None

    if request.method == 'POST':
        form = QRPrinterSelectionForm(request.POST, printers=installed)
        if form.is_valid():
            printer_name = form.cleaned_data['printer_name']

            if printer_name.startswith("‚ùå"):
                if setting:
                    setting.delete()
                messages.success(request, "‚úÖ QR printer removed. A4 layout will be used.")
            else:
                if setting:
                    setting.printer_name = printer_name
                    setting.save()
                else:
                    QRPrinterSetting.objects.create(dealer=dealer, printer_name=printer_name)
                messages.success(request, f"‚úÖ QR Printer set to: {printer_name}")

            return redirect('dealer:select_qr_printer')

    else:
        form = QRPrinterSelectionForm(printers=installed, initial={
            'printer_name': saved_printer or "‚ùå No QR Printer (Use A4 Sheet)"
        })

    return render(request, 'dealer/select_qr_printer.html', {
        'form': form,
        'saved_printer': saved_printer,
    })

def generate_qr_base64(data):
    """Generate QR code and return as base64 string"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=2,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

@login_required
def print_with_qr(request, cnote_number):
    print(f"\nüõ†Ô∏è Starting QR print for CN: {cnote_number}")
    
    cnote = get_object_or_404(CNotes, cnote_number=cnote_number)
    articles = Article.objects.filter(cnote=cnote).order_by('id')
    total_art = sum(article.art for article in articles if article.art is not None)

    print(f"üìä Total articles: {len(articles)}, Total items: {total_art}")

    try:
        setting = QRPrinterSetting.objects.get(dealer=cnote.dealer)
        printer_name = setting.printer_name
        print(f"üñ®Ô∏è Using printer: {printer_name}")
    except QRPrinterSetting.DoesNotExist:
        printer_name = None
        print("üìÑ No printer configured, using A4 layout")

    qr_images = []
    overall_item_counter = 1
    successful_prints = 0
    failed_prints = 0

    for article in articles:
        item_quantity = article.art if article.art else 1
        item_description = article.said_to_contain

        print(f"üì¶ Processing: {item_description} (Qty: {item_quantity})")

        for item_number in range(1, item_quantity + 1):
            item_type_clean = item_description.upper().replace(' ', '').replace('-', '')[:10]
            unique_tracking_id = f"{cnote.cnote_number}-{item_type_clean}-{item_number}"

            booking_date = cnote.created_at.strftime("%d/%m/%y") if cnote.created_at else "N/A"
            qr_string = f"ID:{unique_tracking_id}|CN:{cnote.cnote_number}|TYPE:{item_description}|NUM:{item_number}/{item_quantity}|DATE:{booking_date}|FROM:{cnote.consignor_name}|TO:{cnote.consignee_name}|DEST:{cnote.delivery_destination.destination_name}|COMPANY:Good Way Express"
            qr_img = generate_qr_base64(qr_string)

            # Complete context for grid template
            single_qr_context = {
                'qr': qr_img,
                'tracking_code': unique_tracking_id,
                'cn': cnote.cnote_number,
                'sender': cnote.consignor_name,
                'item': item_description,
                'receiver': cnote.consignee_name,
                'mob': cnote.consignee_mobile,
                'to': cnote.delivery_destination.destination_name,
                'qty': f"{item_number} of {item_quantity}",
                'from_city': cnote.dealer.city,
                'date': booking_date,
                'cnote': cnote,
                'item_description': item_description,
                'item_number': item_number,
                'total_of_this_type': item_quantity,
                'overall_position': overall_item_counter,
                'total_items': total_art,
            }

            qr_images.append(single_qr_context)

            # Print individual QR if printer is configured
            if printer_name and not printer_name.lower().startswith("‚ùå"):
                print(f"üñ®Ô∏è Printing QR {overall_item_counter}/{total_art}: {item_description} ({item_number}/{item_quantity})")
                
                # Complete context for single print template
                template_context = {
                    'qr': qr_img,
                    'tracking_code': unique_tracking_id,
                    'cn': cnote.cnote_number,
                    'sender': cnote.consignor_name,
                    'item': item_description,
                    'receiver': cnote.consignee_name,
                    'mob': cnote.consignee_mobile,
                    'to': cnote.delivery_destination.destination_name,
                    'from_city': cnote.dealer.city,
                    'date': booking_date,
                    
                    # ‚úÖ ALL QUANTITY VARIABLES - ‡§Ø‡§π‡§æ‡§Ç ‡§π‡•à main fix
                    'item_number': item_number,
                    'total_of_this_type': item_quantity,
                    'overall_position': overall_item_counter,
                    'total_items': total_art,
                    
                    # ‚úÖ FALLBACK VARIABLES
                    'current_item': overall_item_counter,
                    'total_quantity': total_art,
                    'position': overall_item_counter,
                    'total': total_art,
                    'qty': f"{item_number} of {item_quantity}",
                }
                
                # Debug print to verify values
                print(f"üìä Context: overall_position={overall_item_counter}, total_items={total_art}, item_number={item_number}, total_of_this_type={item_quantity}")
                
                try:
                    html_output = render_to_string('dealer/qr_single_print_template.html', template_context)
                    success = send_qr_to_printer_using_html(html_output, printer_name)
                    
                    if success:
                        successful_prints += 1
                        print(f"‚úÖ Successfully printed QR {overall_item_counter}")
                    else:
                        failed_prints += 1
                        print(f"‚ùå Failed to print QR {overall_item_counter}")
                        
                except Exception as e:
                    failed_prints += 1
                    print(f"‚ùå Template error for QR {overall_item_counter}: {e}")

            overall_item_counter += 1

    # Final context for templates
    context = {
        'cnote': cnote,
        'dealer': cnote.dealer,
        'articles': articles,
        'total_art': total_art,
        'qr_images': qr_images,
        'company_name': 'Good Way Express',
    }

    # Show results and return appropriate template
    if printer_name and not printer_name.lower().startswith("‚ùå"):
        if successful_prints > 0:
            messages.success(request, f"‚úÖ {successful_prints} QR codes printed successfully!")
        if failed_prints > 0:
            messages.error(request, f"‚ùå {failed_prints} QR codes failed to print.")
        
        print(f"üìä Print Summary: {successful_prints} successful, {failed_prints} failed")
        return render(request, 'dealer/cnote_success_only.html', context)
    else:
        return render(request, 'dealer/cnote_success_with_qr_grid.html', context)  



logger = logging.getLogger(__name__)
def search_cnote(request):
    cnote_number = request.GET.get('cnote_number', '')
    try:
        cnote = CNotes.objects.get(cnote_number=cnote_number)
        return redirect('dealer:view_cnote', cnote_number=cnote_number)
    except CNotes.DoesNotExist:
        logger.warning(f"CNote not found: {cnote_number}")
        messages.error(request, f"CNote with number {cnote_number} not found.")
        return redirect(reverse('dealer:create_cnotes'))
 
logger = logging.getLogger(__name__)
@require_GET
def party_suggestions(request):
    try:
        query = request.GET.get('query', '').strip().upper()
        destination = request.GET.get('destination', '').strip().upper()
        destination_wise = request.GET.get('destination_wise', 'true') == 'true'

        logger.info(f"Party suggestion query: {query}, Destination: {destination}, Filter: {destination_wise}")

        if len(query) < 2:
            return JsonResponse([], safe=False)

        # Base filter for party suggestions
        party_filter = Q(party_name__istartswith=query) | \
                       Q(party_code__istartswith=query) | \
                       Q(mobile_number_1__istartswith=query) | \
                       Q(gst_no__istartswith=query)

        # Apply destination filter if enabled
        if destination and destination_wise:
            party_filter &= Q(city__iexact=destination)  # Exact city match

        # Fetch filtered results
        parties = PartyMaster.objects.filter(party_filter).distinct().order_by('party_name')[:10]

        data = [{
            'party_code': party.party_code,
            'party_name': party.party_name,
            'mobile_number_1': party.mobile_number_1,
            'gst_no': party.gst_no or '',
            'address': party.address,
            'city': party.city,
            'state': party.state
        } for party in parties]

        return JsonResponse(data, safe=False)

    except Exception as e:
        logger.error(f"Error in party_suggestions: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def update_cnote(request, cnote_id):
    try:
        cnote = CNotes.objects.get(id=cnote_id, dealer=request.user.dealer)
        
        if cnote.status != 'booked':
            return JsonResponse({'success': False, 'error': 'Only booked CNotes can be updated.'})

        # Update CNote fields
        cnote.booking_type = request.POST.get('booking_type', cnote.booking_type)
        cnote.delivery_type = request.POST.get('delivery_type', cnote.delivery_type)
        cnote.delivery_method = request.POST.get('delivery_method', cnote.delivery_method)
        
        # Handle delivery_destination
        delivery_destination_id = request.POST.get('delivery_destination')
        if delivery_destination_id:
            cnote.delivery_destination = DeliveryDestination.objects.get(id=delivery_destination_id)
        
        cnote.eway_bill_number = request.POST.get('eway_bill_number', cnote.eway_bill_number)
        cnote.manual_date = request.POST.get('manual_date', cnote.manual_date)
        cnote.manual_cnote_number = request.POST.get('manual_cnote_number', cnote.manual_cnote_number)
        cnote.manual_cnote_type = request.POST.get('manual_cnote_type', cnote.manual_cnote_type)
        cnote.payment_type = request.POST.get('payment_type', cnote.payment_type)
        cnote.consignor_name = request.POST.get('consignor_name', cnote.consignor_name)
        cnote.consignor_mobile = request.POST.get('consignor_mobile', cnote.consignor_mobile)
        cnote.consignor_gst = request.POST.get('consignor_gst', cnote.consignor_gst)
        cnote.consignor_address = request.POST.get('consignor_address', cnote.consignor_address)
        cnote.consignee_name = request.POST.get('consignee_name', cnote.consignee_name)
        cnote.consignee_mobile = request.POST.get('consignee_mobile', cnote.consignee_mobile)
        cnote.consignee_gst = request.POST.get('consignee_gst', cnote.consignee_gst)
        cnote.consignee_address = request.POST.get('consignee_address', cnote.consignee_address)
        cnote.actual_weight = Decimal(request.POST.get('actual_weight', cnote.actual_weight))
        cnote.charged_weight = Decimal(request.POST.get('charged_weight', cnote.charged_weight))
        cnote.weight_rate = Decimal(request.POST.get('weight_rate', cnote.weight_rate))
        cnote.weight_amount = Decimal(request.POST.get('weight_amount', cnote.weight_amount))
        cnote.fix_amount = Decimal(request.POST.get('fix_amount', cnote.fix_amount))
        cnote.invoice_number = request.POST.get('invoice_number', cnote.invoice_number)
        cnote.declared_value = Decimal(request.POST.get('declared_value', cnote.declared_value))
        cnote.risk_type = request.POST.get('risk_type', cnote.risk_type)
        cnote.pod_required = request.POST.get('pod_required', cnote.pod_required)
        
        # Update charges
        cnote.freight = Decimal(request.POST.get('freight', cnote.freight))
        cnote.docket_charge = Decimal(request.POST.get('docket_charge', cnote.docket_charge))
        cnote.door_delivery_charge = Decimal(request.POST.get('door_delivery_charge', cnote.door_delivery_charge))
        cnote.handling_charge = Decimal(request.POST.get('handling_charge', cnote.handling_charge))
        cnote.pickup_charge = Decimal(request.POST.get('pickup_charge', cnote.pickup_charge))
        cnote.transhipment_charge = Decimal(request.POST.get('transhipment_charge', cnote.transhipment_charge))
        cnote.insurance = Decimal(request.POST.get('insurance', cnote.insurance))
        cnote.fuel_surcharge = Decimal(request.POST.get('fuel_surcharge', cnote.fuel_surcharge))
        cnote.commission = Decimal(request.POST.get('commission', cnote.commission))
        cnote.other_charge = Decimal(request.POST.get('other_charge', cnote.other_charge))
        cnote.carrier_risk = Decimal(request.POST.get('carrier_risk', cnote.carrier_risk))
        
        # Recalculate grand total
        cnote.grand_total = (
            cnote.freight +
            cnote.docket_charge +
            cnote.door_delivery_charge +
            cnote.handling_charge +
            cnote.pickup_charge +
            cnote.transhipment_charge +
            cnote.insurance +
            cnote.fuel_surcharge +
            cnote.commission +
            cnote.other_charge +
            cnote.carrier_risk
        )

        cnote.total_art = int(request.POST.get('total_art', cnote.total_art))

        cnote.save()

        # Update articles
        article_types = request.POST.getlist('article_type[]')
        quantities = request.POST.getlist('quantity[]')
        descriptions = request.POST.getlist('description[]')
        amounts = request.POST.getlist('amount[]')
        art_type_ids = request.POST.getlist('art_type[]')

        # Delete existing articles
        cnote.articles.all().delete()

        # Create new articles
        for i in range(len(article_types)):
            art_type = ArtType.objects.get(id=art_type_ids[i]) if art_type_ids[i] else None
            Article.objects.create(
                cnote=cnote,
                article_type=article_types[i],
                art=int(quantities[i]),
                art_type=art_type,
                said_to_contain=descriptions[i],
                art_amount=Decimal(amounts[i])
            )

        return JsonResponse({'success': True})
    except CNotes.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'CNote not found.'})
    except DeliveryDestination.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid delivery destination.'})
    except ArtType.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid article type.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})





logger = logging.getLogger(__name__)

@login_required
def import_cnotes(request):
    # Get the dealer information at the start
    try:
        dealer = request.user.dealer
    except AttributeError:
        messages.error(request, "Dealer information not found.")
        return redirect("some_error_page")

    if request.method == "POST":
        file = request.FILES.get("excel_file")

        if not file:
            messages.error(request, "‚ö†Ô∏è Please select an Excel file to upload.")
            return redirect("dealer:import_cnotes")

        try:
            logger.info("üìÇ Processing Excel File Upload...")
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            failed_rows = []
            success_count = 0
            logger.info(f"‚úÖ Logged-in Dealer: {dealer.name} (ID: {dealer.dealer_id})")

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    logger.info(f"üìå Processing Row {index}: {row}")
                    
                    # Modified data extraction based on new column order
                    booking_type = row[2] or "sundry"
                    delivery_type = row[3] or "economy"
                    delivery_method = row[4] or "DOOR"
                    eway_bill_number = row[5] or "0"
                    payment_type = row[6] or "TBB"
                    consignor_name = row[7] or dealer.name
                    consignor_mobile = row[8] or "0000000000"
                    consignor_gst = row[9] or "UNKNOWN"
                    consignor_address = row[10] or "UNKNOWN"
                    consignee_name = row[11] or "UNKNOWN"
                    delivery_destination_name = row[12] or "UNKNOWN"
                    consignee_address = row[13] or "UNKNOWN"
                    invoice_number = row[14] or ""
                    article_quantity = row[15] or 0
                    consignee_mobile = row[16] or "0000000000"
                    consignee_gst = row[17] or "UNKNOWN"
                    actual_weight = row[18] or 0.0
                    charged_weight = row[19] or actual_weight
                    weight_rate = row[20] or 0.0
                    weight_amount = row[21] or 0.0
                    fix_amount = row[22] or 0.0
                    declared_value = row[23] or 0.0
                    risk_type = row[24] or "owner"
                    pod_required = row[25] or "yes"
                    freight = row[26] or 0.0
                    docket_charge = row[27] or 0.0
                    door_delivery_charge = row[28] or 0.0
                    handling_charge = row[29] or 0.0
                    pickup_charge = row[30] or 0.0
                    transhipment_charge = row[31] or 0.0
                    insurance = row[32] or 0.0
                    fuel_surcharge = row[33] or 0.0
                    commission = row[34] or 0.0
                    other_charge = row[35] or 0.0
                    carrier_risk = row[36] or 0.0
                    grand_total = row[37] or 0.0
                    total_art = row[39] or 0

                    def safe_float(value, default=0.0):
                        try:
                            return float(value) if value not in [None, "", "NULL"] else default
                        except ValueError:
                            return default

                    actual_weight = safe_float(row[18])
                    charged_weight = safe_float(row[19], actual_weight)
                    weight_rate = safe_float(row[20])
                    weight_amount = actual_weight * weight_rate
                    declared_value = safe_float(row[23])
                    total_art = int(row[39]) if row[39] else 0
                    
                    delivery_destination, _ = DeliveryDestination.objects.get_or_create(
                        destination_name=delivery_destination_name
                    )
                    
                    with transaction.atomic():
                        last_cnote = CNotes.objects.filter(dealer=dealer).order_by('-id').first()
                        next_cnote_number = f"{dealer.dealer_code}-{int(last_cnote.cnote_number.split('-')[-1]) + 1:04d}" if last_cnote else f"{dealer.dealer_code}-0001"

                        cnote = CNotes.objects.create(
                            cnote_number=next_cnote_number,
                            dealer=dealer,
                            booking_type=booking_type,
                            delivery_type=delivery_type,
                            delivery_method=delivery_method,
                            delivery_destination=delivery_destination,
                            eway_bill_number=eway_bill_number,
                            payment_type=payment_type,
                            consignor_name=consignor_name,
                            consignor_mobile=consignor_mobile,
                            consignor_address=consignor_address,
                            consignor_gst=consignor_gst,
                            consignee_name=consignee_name,
                            consignee_mobile=consignee_mobile,
                            consignee_gst=consignee_gst,
                            consignee_address=consignee_address,
                            actual_weight=actual_weight,
                            charged_weight=charged_weight,
                            weight_rate=weight_rate,
                            weight_amount=weight_amount,
                            fix_amount=fix_amount,
                            invoice_number=invoice_number,
                            declared_value=declared_value,
                            risk_type=risk_type,
                            pod_required=pod_required,
                            freight=freight,
                            docket_charge=docket_charge,
                            door_delivery_charge=door_delivery_charge,
                            handling_charge=handling_charge,
                            pickup_charge=pickup_charge,
                            transhipment_charge=transhipment_charge,
                            insurance=insurance,
                            fuel_surcharge=fuel_surcharge,
                            commission=commission,
                            other_charge=other_charge,
                            carrier_risk=carrier_risk,
                            grand_total=grand_total,
                            total_art=total_art,
                            status="booked"
                        )
                        
                        def get_art_type(art_type_name):
                            art_type_obj = ArtType.objects.filter(art_type_name=art_type_name).first()
                            return art_type_obj if art_type_obj else ArtType.objects.get(id=1)  # Default ID: 1

                        for i in range(total_art):
                            Article.objects.create(
                                cnote=cnote,
                                article_type=safe_get(row, 40, "UNKNOWN"),
                                art=safe_get(row, 15, "UNKNOWN"),  # Article Quantity is now at index 15
                                art_type=get_art_type(safe_get(row, 41, "UNKNOWN")),
                                said_to_contain=safe_get(row, 42, "UNKNOWN"),
                                art_amount=safe_get(row, 43, 0.0)
                            )

                    success_count += 1
                    logger.info(f"‚úÖ CNote Created Successfully: {next_cnote_number}")
                
                except Exception as e:
                    logger.error(f"‚ùå Error at Row {index}: {str(e)}")
                    failed_rows.append({"row": row, "error": str(e)})

            if success_count > 0:
                messages.success(request, f"‚úÖ {success_count} CNotes imported successfully!")
            if failed_rows:
                messages.error(request, f"‚ö†Ô∏è Some rows failed to upload. Check errors below.")
            return render(request, "dealer/import_cnotes.html", {"failed_rows": failed_rows})
        
        except Exception as e:
            logger.error(f"‚ùå Critical Error: {str(e)}")
            messages.error(request, f"‚ùå Error processing file: {e}")
            return redirect("dealer:import_cnotes")
    
    return render(request, "dealer/import_cnotes.html", {"dealer": dealer})

def safe_get(row, index, default=""):
    try:
        return row[index] if row[index] not in [None, "", "NULL"] else default
    except IndexError:
        return default

    
@login_required
def export_excel_template(request):
    try:
        dealer = Dealer.objects.get(user=request.user)
    except Dealer.DoesNotExist:
        return HttpResponse("Dealer not found", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CNotes & Articles"

    # Modified Headers with new order
    headers = [
        "Dealer ID", "Dealer Name", "Booking Type", "Delivery Type", "Delivery Method", "EWay Bill Number",
        "Payment Type", "Consignor Name", "Consignor Mobile", "Consignor GST", "Consignor Address",
        "Consignee Name", "Delivery Destination", "Consignee Address", "Invoice Number", "Article Quantity",
        "Consignee Mobile", "Consignee GST", "Actual Weight", "Charged Weight", "Weight Rate", "Weight Amount",
        "Fix Amount", "Declared Value", "Risk Type", "POD Required", "Freight", "Docket Charge",
        "Door Delivery Charge", "Handling Charge", "Pickup Charge", "Transhipment Charge", "Insurance",
        "Fuel Surcharge", "Commission", "Other Charge", "Carrier Risk", "Grand Total", "Status",
        "Total Articles", "Article Type", "Art Type", "Said To Contain", "Article Amount"
    ]
    ws.append(headers)

    for _ in range(50):  # 50 Sample rows
        # Modified CNote Sample Data
        cnote_data = [
            dealer.dealer_id, dealer.name, "sundry", "economy", "DOOR", "", "TBB", dealer.name,
            dealer.mobile_number_1, "", dealer.address, "", "", "", "", 0, "", "", 0, 0, 0, 0, 0, 0,
            "owner", "yes", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "booked", 1
        ]
        
        # Modified Article Sample Data
        article_data = ["Article", "Small Box", "Box", 0]

        # Add complete row in a single line
        ws.append(cnote_data + article_data)

    # Generate Excel response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="CNote_Booking_Template.xlsx"'
    wb.save(response)

    return response



@login_required
@require_GET
def get_consignee_data(request):
    """
    Fetch consignee data based on destination.
    """
    try:
        destination = request.GET.get("destination", "").strip().upper()
        destination_wise = request.GET.get("destination_wise", "true") == "true"  # Convert to boolean

        logger.info(f"Fetching consignee data for destination: {destination}, filter: {destination_wise}")

        if not destination:
            return JsonResponse({"success": False, "message": "Destination is required"}, status=400)

        # ‚úÖ FIX: Filter consignees with exact city match
        if destination_wise:
            parties = PartyMaster.objects.filter(city__iexact=destination).distinct().order_by('party_name')
        else:
            parties = PartyMaster.objects.all().distinct().order_by('party_name')

        # ‚úÖ FIX: Ensure only correct destination parties are returned
        consignee_list = [
            {
                "name": party.party_name,
                "code": party.party_code,
                "mobile": party.mobile_number_1,
                "gst": party.gst_no or "",
                "address": party.address,
                "city": party.city,
                "state": party.state,
                "pincode": party.pincode,
                "email": party.email,
                "is_tbb": party.is_tbb
            }
            for party in parties
            if party.city.strip().upper() == destination  # ‚úÖ Final filtering
        ]

        logger.info(f"‚úÖ Found {len(consignee_list)} matching consignees for destination: {destination}")

        return JsonResponse({"success": True, "consignees": consignee_list})

    except Exception as e:
        logger.error(f"‚ùå Error in get_consignee_data: {str(e)}")
        return JsonResponse({"success": False, "message": f"Error fetching consignee data: {str(e)}"}, status=500)


@login_required
def check_loading_sheet(request):
    try:
        dealer = request.user.dealer

        # ‚úÖ Fetch Selected CNotes from Query Params
        selected_cnote_ids = request.GET.getlist("selected_cnotes")

        if not selected_cnote_ids:
            return render(request, "dealer/check_loading_sheet.html", {"error_message": "No CNotes selected."})

        # ‚úÖ Fetch CNotes directly instead of LoadingSheetDetail
        selected_cnotes = CNotes.objects.filter(id__in=selected_cnote_ids, dealer=dealer)

        context = {
            "loading_sheet_details": selected_cnotes,
            "dealer": dealer,
            "current_date": timezone.now(),
        }

        return render(request, "dealer/check_loading_sheet.html", context)

    except Exception as e:
        logger.error(f"Error rendering Check Loading Sheet: {str(e)}")
        return render(request, "dealer/check_loading_sheet.html", {"error_message": "Error loading data."})


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db import transaction
from dealer_app.models import CNotes, Article, DeliveryDestination, Dealer, ArtType
import json
import logging
import re
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from decimal import Decimal

logger = logging.getLogger(__name__)

@login_required
def view_cnote(request, cnote_number):
    try:
        dealer = get_object_or_404(Dealer, user=request.user)
        
        if not cnote_number or not re.match(r'^[A-Z0-9-]{1,50}$', cnote_number):
            logger.warning(f"Invalid cnote_number format: {cnote_number}")
            messages.error(request, "Invalid CNote number format.")
            return redirect('dealer:create_cnotes')
        
        cnote = get_object_or_404(CNotes, cnote_number=cnote_number)
        articles = Article.objects.filter(cnote=cnote).select_related('art_type')
        art_types = ArtType.objects.all()
        
        art_types_json = json.dumps([
            {'id': art_type.id, 'name': art_type.art_type_name}
            for art_type in art_types
        ])
        
        context = {
            'dealer': dealer,
            'cnote': cnote,
            'articles': articles,
            'art_types': art_types,
            'art_types_json': art_types_json,
        }
        
        logger.info(f"Rendering CNote {cnote_number} for dealer {dealer.name}")
        return render(request, 'dealer/view_cnote.html', context)
    
    except ObjectDoesNotExist as e:
        logger.error(f"Object not found for CNote {cnote_number}: {str(e)}")
        messages.error(request, "CNote or dealer not found.")
        return redirect('dealer:create_cnotes')
    except Exception as e:
        logger.error(f"Error rendering CNote {cnote_number}: {str(e)}")
        messages.error(request, "An error occurred while loading the CNote.")
        return redirect('dealer:create_cnotes')
        
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import logging
from dealer_app.models import DeliveryDestination

logger = logging.getLogger(__name__)

@login_required
def fetch_cities(request):
    try:
        query = request.GET.get('q', '')

        # Query all DeliveryDestination entries, excluding problematic ones
        destinations = DeliveryDestination.objects.exclude(
            destination_name__in=['UNKNOWN', 'Address']
        )

        # Apply search query if provided
        if query:
            destinations = destinations.filter(destination_name__icontains=query)

        # Get id and destination_name
        cities = destinations.values('id', 'destination_name')
        cities_list = list(cities)

        # Deduplicate based on destination_name (case-insensitive)
        seen_names = set()
        deduplicated_cities = []
        for city in cities_list:
            name_lower = city['destination_name'].lower().strip()  # Normalize by stripping spaces
            if name_lower not in seen_names:
                seen_names.add(name_lower)
                deduplicated_cities.append(city)

        # Log and return the results
        user_identifier = request.user.get_username() or str(request.user.id)
        logger.info(f"Fetched {len(deduplicated_cities)} destinations for user {user_identifier}")
        return JsonResponse({'success': True, 'cities': deduplicated_cities})

    except Exception as e:
        logger.error(f"Error fetching cities: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to fetch cities'}, status=500)

@login_required
def update_freight(request):
    if request.method != 'POST':
        logger.warning(f"Invalid request method for update_freight: {request.method}")
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        cnote_number = data.get('cnote_number')
        new_freight = data.get('freight')
        
        if not cnote_number or not isinstance(cnote_number, str):
            logger.warning("Invalid or missing cnote_number")
            return JsonResponse({'success': False, 'error': 'Invalid CNote number'}, status=400)
        
        if not re.match(r'^[A-Z0-9-]{1,50}$', cnote_number):
            logger.warning(f"Invalid cnote_number format: {cnote_number}")
            return JsonResponse({'success': False, 'error': 'Invalid CNote number format'}, status=400)
        
        if new_freight is None:
            logger.warning(f"Missing freight value for CNote {cnote_number}")
            return JsonResponse({'success': False, 'error': 'Freight value is required'}, status=400)
        
        try:
            new_freight = Decimal(str(new_freight))
            if new_freight < 0:
                logger.warning(f"Negative freight value for CNote {cnote_number}: {new_freight}")
                return JsonResponse({'success': False, 'error': 'Freight cannot be negative'}, status=400)
            if new_freight > 1000000:
                logger.warning(f"Excessive freight value for CNote {cnote_number}: {new_freight}")
                return JsonResponse({'success': False, 'error': 'Freight value too large'}, status=400)
        except (ValueError, TypeError):
            logger.warning(f"Invalid freight value for CNote {cnote_number}: {new_freight}")
            return JsonResponse({'success': False, 'error': 'Invalid freight value'}, status=400)
        
        with transaction.atomic():
            cnote = get_object_or_404(CNotes, cnote_number=cnote_number)
            cnote.freight = new_freight
            cnote.grand_total = (
                Decimal(str(cnote.freight or 0)) +
                Decimal(str(cnote.docket_charge or 0)) +
                Decimal(str(cnote.door_delivery_charge or 0)) +
                Decimal(str(cnote.handling_charge or 0)) +
                Decimal(str(cnote.pickup_charge or 0)) +
                Decimal(str(cnote.transhipment_charge or 0)) +
                Decimal(str(cnote.insurance or 0)) +
                Decimal(str(cnote.fuel_surcharge or 0)) +
                Decimal(str(cnote.commission or 0)) +
                Decimal(str(cnote.other_charge or 0)) +
                Decimal(str(cnote.carrier_risk or 0))
            )
            cnote.save()
            
            logger.info(f"Freight updated for CNote {cnote_number} to {new_freight}")
            return JsonResponse({
                'success': True,
                'freight': float(new_freight),
                'grand_total': float(cnote.grand_total)
            })
    
    except CNotes.DoesNotExist:
        logger.warning(f"CNote {cnote_number} not found")
        return JsonResponse({'success': False, 'error': 'CNote not found'}, status=404)
    except Exception as e:
        logger.error(f"Error updating freight for CNote {cnote_number}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to update freight'}, status=400)

@login_required
def save_cnote(request):
    if request.method != 'POST':
        logger.warning(f"Invalid request method for save_cnote: {request.method}")
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        cnote_number = data.get('cnote_number')
        actual_weight = data.get('actual_weight')
        charged_weight = data.get('charged_weight')
        delivery_destination_id = data.get('delivery_destination_id')
        articles = data.get('articles', [])
        charges = data.get('charges', {})

        if not cnote_number or not isinstance(cnote_number, str):
            logger.warning("Invalid or missing cnote_number")
            return JsonResponse({'success': False, 'error': 'Invalid CNote number'}, status=400)
        
        if not re.match(r'^[A-Z0-9-]{1,50}$', cnote_number):
            logger.warning(f"Invalid cnote_number format: {cnote_number}")
            return JsonResponse({'success': False, 'error': 'Invalid CNote number format'}, status=400)
        
        if actual_weight is None or charged_weight is None:
            logger.warning(f"Missing weight values for CNote {cnote_number}")
            return JsonResponse({'success': False, 'error': 'Weight values are required'}, status=400)
        
        try:
            actual_weight = Decimal(str(actual_weight))
            charged_weight = Decimal(str(charged_weight))
            if actual_weight < 0 or charged_weight < 0:
                logger.warning(f"Negative weight values for CNote {cnote_number}")
                return JsonResponse({'success': False, 'error': 'Weights cannot be negative'}, status=400)
        except (ValueError, TypeError):
            logger.warning(f"Invalid weight values for CNote {cnote_number}")
            return JsonResponse({'success': False, 'error': 'Invalid weight values'}, status=400)
        
        if not delivery_destination_id:
            logger.warning(f"Missing delivery_destination_id for CNote {cnote_number}")
            return JsonResponse({'success': False, 'error': 'Delivery destination is required'}, status=400)
        
        if not articles:
            logger.warning(f"No articles provided for CNote {cnote_number}")
            return JsonResponse({'success': False, 'error': 'At least one article is required'}, status=400)

        with transaction.atomic():
            cnote = get_object_or_404(CNotes, cnote_number=cnote_number)
            cnote.actual_weight = actual_weight
            cnote.charged_weight = charged_weight
            cnote.delivery_destination = get_object_or_404(DeliveryDestination, id=delivery_destination_id)
            
            for field in ['docket_charge', 'door_delivery_charge', 'handling_charge', 'pickup_charge',
                         'transhipment_charge', 'insurance', 'fuel_surcharge', 'commission',
                         'other_charge', 'carrier_risk']:
                value = charges.get(field, 0)
                try:
                    value = Decimal(str(value)) if value is not None else Decimal('0')
                    if value < 0:
                        logger.warning(f"Negative {field} for CNote {cnote_number}: {value}")
                        return JsonResponse({'success': False, 'error': f'{field.replace("_", " ").title()} cannot be negative'}, status=400)
                    setattr(cnote, field, value)
                except (ValueError, TypeError):
                    logger.warning(f"Invalid {field} for CNote {cnote_number}: {value}")
                    return JsonResponse({'success': False, 'error': f'Invalid {field.replace("_", " ").title()} value'}, status=400)
            
            cnote.grand_total = (
                Decimal(str(cnote.freight or 0)) +
                Decimal(str(cnote.docket_charge or 0)) +
                Decimal(str(cnote.door_delivery_charge or 0)) +
                Decimal(str(cnote.handling_charge or 0)) +
                Decimal(str(cnote.pickup_charge or 0)) +
                Decimal(str(cnote.transhipment_charge or 0)) +
                Decimal(str(cnote.insurance or 0)) +
                Decimal(str(cnote.fuel_surcharge or 0)) +
                Decimal(str(cnote.commission or 0)) +
                Decimal(str(cnote.other_charge or 0)) +
                Decimal(str(cnote.carrier_risk or 0))
            )
            cnote.save()

            existing_article_ids = set(Article.objects.filter(cnote=cnote).values_list('id', flat=True))
            new_article_ids = set(int(article.get('id')) for article in articles if article.get('id') and article.get('id').isdigit())
            articles_to_delete = existing_article_ids - new_article_ids
            Article.objects.filter(id__in=articles_to_delete).delete()

            for article_data in articles:
                art_type_id = article_data.get('art_type')
                quantity = article_data.get('quantity')
                description = article_data.get('description')
                amount = article_data.get('amount')
                article_id = article_data.get('id')

                if not all([art_type_id, quantity is not None, description, amount is not None]):
                    logger.warning(f"Incomplete article data for CNote {cnote_number}")
                    return JsonResponse({'success': False, 'error': 'Incomplete article data'}, status=400)
                
                try:
                    art_type = ArtType.objects.get(id=art_type_id)
                except ArtType.DoesNotExist:
                    logger.warning(f"Invalid art_type_id {art_type_id} for CNote {cnote_number}")
                    return JsonResponse({'success': False, 'error': 'Invalid article type'}, status=400)
                
                try:
                    quantity = int(quantity)
                    amount = Decimal(str(amount))
                    if quantity <= 0 or amount < 0:
                        logger.warning(f"Invalid article values for CNote {cnote_number}: quantity={quantity}, amount={amount}")
                        return JsonResponse({'success': False, 'error': 'Article quantity must be positive and amount cannot be negative'}, status=400)
                except (ValueError, TypeError):
                    logger.warning(f"Invalid article values for CNote {cnote_number}")
                    return JsonResponse({'success': False, 'error': 'Invalid article quantity or amount'}, status=400)

                if article_id and article_id.isdigit():
                    try:
                        article = Article.objects.get(id=int(article_id), cnote=cnote)
                        article.art_type = art_type
                        article.art = quantity
                        article.said_to_contain = description
                        article.art_amount = amount
                        article.save()
                    except Article.DoesNotExist:
                        logger.warning(f"Article {article_id} not found for CNote {cnote_number}, creating new")
                        Article.objects.create(
                            cnote=cnote,
                            art_type=art_type,
                            art=quantity,
                            said_to_contain=description,
                            art_amount=amount
                        )
                else:
                    Article.objects.create(
                        cnote=cnote,
                        art_type=art_type,
                        art=quantity,
                        said_to_contain=description,
                        art_amount=amount
                    )
            
            logger.info(f"CNote {cnote_number} updated successfully")
            return JsonResponse({'success': True})
    
    except CNotes.DoesNotExist:
        logger.warning(f"CNote {cnote_number} not found")
        return JsonResponse({'success': False, 'error': 'CNote not found'}, status=404)
    except DeliveryDestination.DoesNotExist:
        logger.warning(f"Delivery destination {delivery_destination_id} not found for CNote {cnote_number}")
        return JsonResponse({'success': False, 'error': 'Invalid delivery destination'}, status=400)
    except Exception as e:
        logger.error(f"Error updating CNote {cnote_number}: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Failed to update CNote'}, status=400)
